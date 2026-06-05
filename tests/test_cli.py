from pathlib import Path

from openpyxl import Workbook, load_workbook

import wlc_role_acl_collector.cli as cli
from wlc_role_acl_collector.cli import main
from wlc_role_acl_collector.interactive import prompt_controller_targets
from wlc_role_acl_collector.models import CollectionResult


def test_cli_collect_offline(tmp_path):
    controllers = tmp_path / "controllers.csv"
    controllers.write_text(
        "name,host,protocol,port,device_type,username_env,password_env,enable_password_env\n"
        "sample_controller,192.0.2.10,ssh,22,aruba_os,,,\n",
        encoding="utf-8",
    )
    fixture_root = Path(__file__).parent / "fixtures"
    role_networks = tmp_path / "role_networks.xlsx"
    _write_role_networks(role_networks)

    exit_code = main(
        [
            "collect",
            "--controllers",
            str(controllers),
            "--offline-raw-dir",
            str(fixture_root),
            "--output-dir",
            str(tmp_path / "outputs"),
            "--role-networks",
            str(role_networks),
        ]
    )

    assert exit_code == 0
    reports = list((tmp_path / "outputs").glob("*/ssid_role_acl_report.xlsx"))
    assert len(reports) == 1
    workbook = load_workbook(reports[0], read_only=True)
    assert "Local_Role_Networks" in workbook.sheetnames


def test_cli_interactive_defaults_to_ssh(monkeypatch, tmp_path):
    prompts = iter(
        [
            "10.10.10.10",
            "",
            "",
            "",
            "admin",
            "",
        ]
    )
    captured = {}

    monkeypatch.setattr(
        cli,
        "prompt_controller_targets",
        lambda: prompt_controller_targets(
            input_func=lambda _prompt: next(prompts),
            password_func=lambda _prompt: "secret",
        ),
    )

    def fake_collect(controller, *, timeout, credentials):
        captured["controller"] = controller
        captured["credentials"] = credentials
        return CollectionResult(controller=controller)

    monkeypatch.setattr(cli, "collect_from_controller", fake_collect)

    exit_code = main(["collect", "--output-dir", str(tmp_path / "outputs")])

    assert exit_code == 0
    assert captured["controller"].host == "10.10.10.10"
    assert captured["controller"].name == "wlc-10.10.10.10"
    assert captured["controller"].protocol == "ssh"
    assert captured["controller"].port == 22
    assert captured["controller"].device_type == "aruba_os"
    assert captured["credentials"].username == "admin"


def _write_role_networks(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["role", "network", "subnet_mask"])
    worksheet.append(["guest-logon", "10.30.0.0", "255.255.255.0"])
    workbook.save(path)
