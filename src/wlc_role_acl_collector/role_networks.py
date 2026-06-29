"""Load optional local Role-to-network Excel mappings.

These mappings are treated as sensitive internal data. By default they are read
for the current run only and are not exported into generated reports.
"""

from __future__ import annotations

import ipaddress
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import RoleNetworkDefinition


class RoleNetworkDefinitionError(ValueError):
    pass


@dataclass(frozen=True)
class RoleNetworkLoadSummary:
    definitions: list[RoleNetworkDefinition]
    role_count: int
    network_count: int
    duplicate_count: int
    source_file: str


_EXCEL_FORMAT_HINT = (
    "Use config\\role_networks.example.xlsx as the template, edit only the rows, "
    "then save it as Excel Workbook (*.xlsx). Do not rename a CSV, HTML, or .xls file to .xlsx."
)


_HEADER_ALIASES = {
    "role": {
        "role",
        "rolename",
        "role name",
        "role이름",
        "role 이름",
        "역할",
        "정책이름",
        "정책 이름",
    },
    "network": {
        "network",
        "networkaddress",
        "network address",
        "networkcidr",
        "network cidr",
        "cidr",
        "subnet",
        "네트워크대역",
        "네트워크 대역",
        "대역",
    },
    "subnet_mask": {
        "subnetmask",
        "subnet mask",
        "netmask",
        "mask",
        "서브넷마스크",
        "서브넷 마스크",
    },
}


def load_role_network_definitions(path: Path | str | None) -> list[RoleNetworkDefinition]:
    return load_role_network_definitions_with_summary(path).definitions


def load_role_network_definitions_with_summary(path: Path | str | None) -> RoleNetworkLoadSummary:
    if path is None:
        return RoleNetworkLoadSummary([], role_count=0, network_count=0, duplicate_count=0, source_file="")

    source = Path(path)
    _validate_excel_file(source)

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except zipfile.BadZipFile as exc:
        raise RoleNetworkDefinitionError(_not_real_xlsx_message(source)) from exc
    except Exception as exc:
        raise RoleNetworkDefinitionError(
            f"Unable to open Role network Excel file: {exc}. {_EXCEL_FORMAT_HINT}"
        ) from exc

    try:
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_index = _find_header_row(rows)
    if header_index is None:
        raise RoleNetworkDefinitionError("Role network Excel file does not contain a header row.")

    columns = _resolve_columns(rows[header_index])
    errors: list[str] = []
    definitions: list[RoleNetworkDefinition] = []
    seen: set[tuple[str, str]] = set()
    duplicate_count = 0

    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if _row_is_blank(row):
            continue

        role = _cell_text(_cell_at(row, columns["role"]))
        network_text = _cell_text(_cell_at(row, columns["network"]))
        mask_column = columns.get("subnet_mask")
        mask_text = _cell_text(_cell_at(row, mask_column)) if mask_column is not None else ""

        if not role:
            errors.append(f"row {row_number}: Role name is required.")
            continue
        if not network_text:
            errors.append(f"row {row_number}: Network is required for role {role}.")
            continue

        try:
            cidr, subnet_mask = _normalize_network(network_text, mask_text)
        except RoleNetworkDefinitionError as exc:
            errors.append(f"row {row_number}: {exc}")
            continue

        dedupe_key = (role.casefold(), cidr)
        if dedupe_key in seen:
            duplicate_count += 1
            continue
        seen.add(dedupe_key)
        definitions.append(
            RoleNetworkDefinition(
                role=role,
                network=cidr,
                subnet_mask=subnet_mask,
                source_file=str(source),
                source_row=row_number,
            )
        )

    if errors:
        suffix = "" if len(errors) <= 8 else f" ... and {len(errors) - 8} more error(s)"
        raise RoleNetworkDefinitionError("; ".join(errors[:8]) + suffix)
    if not definitions:
        raise RoleNetworkDefinitionError("Role network Excel file does not contain any mapping rows.")
    return RoleNetworkLoadSummary(
        definitions=definitions,
        role_count=len({definition.role.casefold() for definition in definitions}),
        network_count=len(definitions),
        duplicate_count=duplicate_count,
        source_file=str(source),
    )


def _validate_excel_file(source: Path) -> None:
    if not source.exists():
        raise RoleNetworkDefinitionError(f"Role network Excel file was not found: {source}")
    if source.name.startswith("~$"):
        raise RoleNetworkDefinitionError(
            "The selected file looks like an Excel temporary lock file. "
            "Close Excel or select the real workbook file that does not start with '~$'."
        )

    suffix = source.suffix.lower()
    if suffix == ".xls":
        raise RoleNetworkDefinitionError(
            "The selected file is an old .xls workbook. "
            "Open it in Excel and save it as Excel Workbook (*.xlsx), then select the new file."
        )
    if suffix not in {".xlsx", ".xlsm"}:
        raise RoleNetworkDefinitionError(
            f"Role network file must be an .xlsx or .xlsm file. {_EXCEL_FORMAT_HINT}"
        )
    if not zipfile.is_zipfile(source):
        raise RoleNetworkDefinitionError(_not_real_xlsx_message(source))


def _not_real_xlsx_message(source: Path) -> str:
    return (
        f"The selected file is not a valid Excel xlsx/xlsm workbook: {source}. "
        f"{_EXCEL_FORMAT_HINT}"
    )


def _find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        if not _row_is_blank(row):
            return index
    return None


def _resolve_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    normalized_headers = {_normalize_header(value): index for index, value in enumerate(header_row)}
    columns: dict[str, int] = {}
    missing: list[str] = []
    for column_name, aliases in _HEADER_ALIASES.items():
        match = next(
            (
                normalized_headers[_normalize_header(alias)]
                for alias in aliases
                if _normalize_header(alias) in normalized_headers
            ),
            None,
        )
        if match is None:
            if column_name != "subnet_mask":
                missing.append(column_name)
        else:
            columns[column_name] = match
    if missing:
        raise RoleNetworkDefinitionError(
            "Role network Excel file is missing required column(s): " + ", ".join(missing)
        )
    return columns


def _normalize_network(network_text: str, mask_text: str) -> tuple[str, str]:
    try:
        if "/" in network_text:
            network = ipaddress.ip_network(network_text, strict=False)
            if mask_text:
                mask_network = ipaddress.ip_network(f"{network.network_address}/{mask_text}", strict=False)
                if mask_network.prefixlen != network.prefixlen:
                    raise RoleNetworkDefinitionError(
                        f"CIDR prefix and subnet mask do not match for {network_text} / {mask_text}."
                    )
        else:
            if not mask_text:
                raise RoleNetworkDefinitionError(
                    f"Subnet mask is required when network is not CIDR: {network_text}."
                )
            network = ipaddress.ip_network(f"{network_text}/{mask_text}", strict=False)
    except RoleNetworkDefinitionError:
        raise
    except ValueError as exc:
        raise RoleNetworkDefinitionError(f"Invalid network or subnet mask: {network_text} {mask_text}".strip()) from exc

    if network.version != 4:
        raise RoleNetworkDefinitionError(f"Only IPv4 networks are supported: {network_text}.")
    return f"{network.network_address}/{network.prefixlen}", str(network.netmask)


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-]+", "", _cell_text(value).casefold())


def _cell_at(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return ""
    return row[index]


def _row_is_blank(row: tuple[Any, ...]) -> bool:
    return all(not _cell_text(value) for value in row)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
