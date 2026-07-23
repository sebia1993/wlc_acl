from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from wlc_role_acl_collector.collector import collect_from_offline_raw
from wlc_role_acl_collector.models import Controller, RoleNetworkDefinition
from wlc_role_acl_collector.report import (
    _acl_rows_html,
    _html2canvas_source,
    _is_role_related_acl,
    _role_image_export_script,
    _write_html,
    build_parsed_controllers,
    create_run_dir,
    write_raw_result,
    write_reports,
)


def test_create_run_dir_adds_safe_label_and_collision_suffix(monkeypatch, tmp_path):
    import wlc_role_acl_collector.report as report

    monkeypatch.setattr(report, "timestamp_slug", lambda: "20260630_120000_000001")

    first = create_run_dir(tmp_path, label="inside/wlc")
    second = create_run_dir(tmp_path, label="inside/wlc")

    assert first.name == "20260630_120000_000001_inside_wlc"
    assert second.name == "20260630_120000_000001_inside_wlc_001"
    assert first.is_dir()
    assert second.is_dir()


def test_write_excel_and_html_report(tmp_path):
    fixture_root = Path(__file__).parent / "fixtures"
    controller = Controller(name="sample_controller", host="192.0.2.10")
    result = collect_from_offline_raw(controller, fixture_root)
    write_raw_result(result, tmp_path / "raw")
    parsed = build_parsed_controllers([result])

    files = write_reports(parsed_controllers=parsed, collection_results=[result], output_dir=tmp_path)

    assert files["xlsx"].exists()
    assert files["html"].exists()
    workbook = load_workbook(files["xlsx"], read_only=True)
    assert {
        "Overview",
        "SSID_Role_Map",
        "Role_Network_Context",
        "Role_ACL_Detail",
        "Alias_Detail",
        "Unresolved",
        "Raw_Commands",
    }.issubset(set(workbook.sheetnames))
    assert "Local_Role_Networks" not in workbook.sheetnames
    acl_headers = [cell.value for cell in next(workbook["Role_ACL_Detail"].iter_rows(max_row=1))]
    assert "source_detail" not in acl_headers
    assert "destination_detail" not in acl_headers
    assert "source_interpretation" in acl_headers
    assert "destination_interpretation" in acl_headers
    assert "role_user_network" in acl_headers
    overview_headers = [cell.value for cell in next(workbook["Overview"].iter_rows(max_row=1))]
    ssid_headers = [cell.value for cell in next(workbook["SSID_Role_Map"].iter_rows(max_row=1))]
    raw_headers = [cell.value for cell in next(workbook["Raw_Commands"].iter_rows(max_row=1))]
    for headers in (overview_headers, ssid_headers, raw_headers):
        assert "site" not in headers
        assert "zone" not in headers
    assert "effective_vlan" in ssid_headers
    assert "role_user_network" in ssid_headers
    assert "network_evidence" in ssid_headers
    assert "network_confidence" in ssid_headers
    assert "assignment_source" in ssid_headers
    assert "configured_vlan" in ssid_headers
    assert "configured_subnet" in ssid_headers
    assert "observed_user_count" in ssid_headers
    context_headers = [cell.value for cell in next(workbook["Role_Network_Context"].iter_rows(max_row=1))]
    assert "role_user_network" in context_headers
    assert "network_confidence" in context_headers
    assert "assignment_source" in context_headers
    assert "configured_vlan" in context_headers
    assert "configured_subnet" in context_headers
    assert "observed_networks" in context_headers
    html = files["html"].read_text(encoding="utf-8")
    assert "corp-employee" in html
    assert "<th>Zone</th>" not in html
    assert 'id="filter"' not in html
    assert 'id="ssid-table"' not in html
    assert 'class="toolbar no-print"' not in html
    assert "<details" not in html
    assert "<th>SSID</th>" not in html
    assert "<th>AP Group</th>" not in html
    assert "<th>AAA Profile</th>" not in html
    assert "<th>Role Type</th>" not in html
    assert "<th>VAP VLAN</th>" not in html
    assert "<th>Effective VLAN</th>" not in html
    assert "Role User Network" not in html
    assert "Local Role Network" not in html
    assert "내부망 전용 보고서" not in html
    assert "Configured Subnet" not in html
    assert "network-context" not in html
    assert "network-chip" not in html
    assert "network-confidence" not in html
    assert 'class="role-tabs no-print"' in html
    assert 'class="role-tab"' in html
    assert 'role="tab"' in html
    assert 'data-role="guest-logon" data-panel-id="role-panel-1"' in html
    assert 'id="role-panel-1" data-role="guest-logon"' in html
    assert 'aria-selected="true"' in html
    assert "afterprint" in html
    assert html.index('<span class="role-tab-name">guest-logon</span>') < html.index(
        '<span class="role-tab-name">corp-employee</span>'
    )
    assert "User IP (current client assigned to this role)" in html
    assert "Any (0.0.0.0/0)" in html
    assert "Source Detail" not in html
    assert "Destination Detail" not in html
    assert "Alias Detail" not in html
    assert 'class="alias-link"' in html
    assert 'class="alias-detail-row" hidden' in html
    assert "<th>Comment</th>" in html
    assert 'class="comment-input"' in html
    assert 'class="comment-status"' in html
    assert 'class="comment-print"' in html
    assert 'id="acl-comments-data"' in html
    assert "localStorage" in html
    assert "commentStorageKey" in html
    assert "임시저장 복원됨" in html
    assert "저장됨" in html
    assert "주석 포함 HTML 저장" in html
    assert "PDF 저장/인쇄" in html
    assert "window.print()" in html
    assert "beforeprint" in html
    assert "@media print" in html
    assert 'class="report-actions no-print"' in html
    assert 'id="save-role-png"' in html
    assert "선택 Role PNG 저장" in html
    assert 'id="role-image-status"' in html
    assert 'class="role-report-description"' in html
    assert 'class="role-description-input"' in html
    assert 'id="role-descriptions-data"' in html
    assert "wlc-role-report-descriptions:" in html
    assert "collectRoleDescriptions()" in html
    assert "syncRoleDescriptionDomValues()" in html
    assert "입력된 설명이 없습니다." in html
    assert "ROLE_IMAGE_MAX_SINGLE_HEIGHT = 12000" in html
    assert "ROLE_IMAGE_PAGE_HEIGHT = 5000" in html
    assert "prepareRoleImageClone" in html
    assert "pruneHiddenRoleRows" in html
    assert "_part_${part}_of_${total}.png" in html
    assert "typeof html2canvas !== 'function'" in html
    assert len(_html2canvas_source()) > 190_000
    assert "html2canvas" in html
    assert "<script src=" not in html
    assert 'id="toggle-raw"' in html
    assert 'aria-pressed="false">Raw 보기</button>' in html
    assert "Raw 숨김" in html
    assert '<th class="raw-column">Raw</th>' in html
    assert 'class="raw raw-column"' in html
    assert ".raw-column { display: none; }" in html
    assert "body.raw-visible .raw-column { display: table-cell; }" in html
    assert "document.body.classList.toggle('raw-visible')" in html
    assert "syncRawToggleButton" in html
    assert "report-header-inner" in html
    assert "report-summary-pill" in html
    assert 'class="executive-summary"' in html
    assert "결론 요약" in html
    assert "확인 필요" in html
    assert "사용자 많은 Role TOP 3" in html
    assert "Access Check 판정 제한 있음" in html
    assert html.index("결론 요약") < html.index('class="report-actions no-print"')
    assert html.index("결론 요약") < html.index('class="access-check no-print"')
    assert html.index("Role ACL Detail") < html.index('class="access-check no-print"')
    assert "rule-badge action-deny" in html
    assert "rule-badge action-special" in html
    assert "rule-badge service-badge" in html
    assert 'class="access-check no-print"' in html
    assert 'id="access-check-data" type="application/json"' in html
    assert 'id="access-check-source"' in html
    assert 'id="access-check-destination"' in html
    assert "자동 - Source/Destination 기준" in html
    assert "Service 자동 모드가" in html
    assert "일치하는 Role ACL 없음" in html
    assert "Access Check는 선택한 Role 이름과 정확히 같은 ACL만 판정합니다." in html
    assert ">검사</button>" in html
    assert "검사 결과 없음." in html
    assert "보고서에 포함된 ACL/Alias 데이터만 사용하며" in html
    assert 'id="access-check-history-data"' not in html
    assert 'id="access-check-history"' not in html
    assert 'id="clear-access-history"' not in html
    assert "rule.sourceMatchers" in html
    assert "runAccessCheck" in html
    assert "accessHighlightRule" in html
    assert "function syncAccessRoleSelection(roleName)" in html
    assert "syncAccessRoleSelection(selectedPanel.dataset.role || '')" in html
    assert "const accessHistoryEnabled = false" in html
    assert "syncAccessHistoryDomValues" in html
    assert "Access Check History" not in html
    assert 'data-rule-id="access-rule-guest-logon-1"' in html
    assert 'class="acl-filter-button toggle-other-acls"' in html
    assert 'data-other-acl-count="5"' in html
    assert "5 other hidden" in html
    assert "Show other ACLs" in html
    assert "Hide other ACLs" in html
    assert 'class="acl-rule-row other-acl-rule" data-other-acl="true" hidden' in html
    assert "syncOtherAclRows" in html
    assert "panel.querySelectorAll('.other-acl-rule')" in html
    assert "document.querySelectorAll('.other-acl-rule')" in html
    assert 'colspan="8"' in html
    assert "alias-type-host" in html
    assert "alias-type-network" in html
    assert "10.10.20.0 255.255.255.0" in html
    raw_text = result.raw_file.read_text(encoding="utf-8")
    assert "[show user-table output redacted]" in raw_text
    assert "corp-user-2" not in raw_text
    assert "aa:bb:cc" not in raw_text


def test_role_image_export_script_preserves_visible_state_and_splits_safely():
    script = _role_image_export_script()

    assert "panel.cloneNode(true)" in script
    assert "row.hidden || style.display === 'none'" in script
    assert "comment-export-text" in script
    assert "role-description-export-text" in script
    assert "ROLE_IMAGE_SCALE = 2" in script
    assert "ROLE_IMAGE_MIN_WIDTH = 1200" in script
    assert "ROLE_IMAGE_MAX_WIDTH = 1800" in script
    assert "assignRoleImageGroups" in script
    assert "roleImageChunks" in script
    assert "downloadRoleCanvas" in script
    assert "roleImageButton.addEventListener('click', saveSelectedRolePng)" in script


def test_write_report_does_not_export_local_role_network_mapping_by_default(tmp_path):
    fixture_root = Path(__file__).parent / "fixtures"
    controller = Controller(name="sample_controller", host="192.0.2.10")
    result = collect_from_offline_raw(controller, fixture_root)
    parsed = build_parsed_controllers([result])
    local_role_networks = [
        RoleNetworkDefinition(
            role="guest-logon",
            network="10.30.0.0/24",
            subnet_mask="255.255.255.0",
            source_file="role_networks.xlsx",
            source_row=2,
        ),
        RoleNetworkDefinition(
            role="guest-logon",
            network="10.31.0.0/24",
            subnet_mask="255.255.255.0",
            source_file="role_networks.xlsx",
            source_row=3,
        ),
    ]

    files = write_reports(
        parsed_controllers=parsed,
        collection_results=[result],
        output_dir=tmp_path,
        local_role_networks=local_role_networks,
    )

    workbook = load_workbook(files["xlsx"], read_only=True)
    assert "Local_Role_Networks" not in workbook.sheetnames
    acl_headers = [cell.value for cell in next(workbook["Role_ACL_Detail"].iter_rows(max_row=1))]
    assert "local_role_networks" not in acl_headers
    assert "local_network_status" not in acl_headers
    assert "local_network_notes" not in acl_headers
    html = files["html"].read_text(encoding="utf-8")
    assert "Local Role Network" not in html
    assert "10.30.0.0/24" not in html
    assert "10.31.0.0/24" not in html
    assert '"localNetworks"' not in html


def test_write_report_exports_local_role_network_mapping_when_enabled(tmp_path):
    fixture_root = Path(__file__).parent / "fixtures"
    controller = Controller(name="sample_controller", host="192.0.2.10")
    result = collect_from_offline_raw(controller, fixture_root)
    parsed = build_parsed_controllers([result])
    local_role_networks = [
        RoleNetworkDefinition(
            role="guest-logon",
            network="10.30.0.0/24",
            subnet_mask="255.255.255.0",
            source_file="role_networks.xlsx",
            source_row=2,
        ),
        RoleNetworkDefinition(
            role="guest-logon",
            network="10.31.0.0/24",
            subnet_mask="255.255.255.0",
            source_file="role_networks.xlsx",
            source_row=3,
        ),
    ]

    files = write_reports(
        parsed_controllers=parsed,
        collection_results=[result],
        output_dir=tmp_path,
        local_role_networks=local_role_networks,
        export_local_role_networks=True,
    )

    workbook = load_workbook(files["xlsx"], read_only=True)
    acl_headers = [cell.value for cell in next(workbook["Role_ACL_Detail"].iter_rows(max_row=1))]
    assert "local_role_networks" in acl_headers
    assert "local_network_status" in acl_headers
    assert "local_network_notes" in acl_headers
    local_rows = [
        {header: value for header, value in zip(
            [cell.value for cell in next(workbook["Local_Role_Networks"].iter_rows(max_row=1))],
            row,
        )}
        for row in workbook["Local_Role_Networks"].iter_rows(min_row=2, values_only=True)
    ]
    assert any(row["role"] == "guest-logon" and row["status"] == "Mismatch" for row in local_rows)
    assert any(row["role"] == "corp-employee" and row["status"] == "Local mapping missing" for row in local_rows)

    html = files["html"].read_text(encoding="utf-8")
    assert "내부망 전용 보고서" in html
    assert "사내 Role 대역표가 포함되어" in html
    assert "Source IP가 해당 Role 대역 밖이면 경고" in html
    assert "Local Role Network" in html
    assert "10.30.0.0/24" in html
    assert "10.31.0.0/24" in html
    assert "Mismatch" in html
    assert "Local mapping missing" in html
    assert "Local Role Network: 10.30.0.0/24, 10.31.0.0/24" in html


def test_acl_rows_keep_collection_order_and_hide_other_acls():
    rows = [
        {"acl": "shared-acl", "sequence": 1},
        {"acl": "guest-logon-acl", "sequence": 2},
        {"acl": "tail-acl", "sequence": 4},
    ]

    html = _acl_rows_html("guest-logon", rows, {})

    assert html.index("<td>shared-acl</td>") < html.index("<td>guest-logon-acl</td>")
    assert html.index("<td>guest-logon-acl</td>") < html.index("<td>tail-acl</td>")
    assert html.count('class="acl-rule-row other-acl-rule" data-other-acl="true" hidden') == 2
    assert html.count('class="acl-rule-row"') == 1


def test_role_related_acl_allows_role_name_inside_acl_name():
    assert _is_role_related_acl("guest-logon", "guest-logon-acl")
    assert _is_role_related_acl("guest-logon", "branch-guest-logon-policy")
    assert not _is_role_related_acl("corp-employee", "corp-acl")


def test_html_hides_zero_user_roles_when_user_table_is_reliable(tmp_path):
    html_path = tmp_path / "report.html"
    _write_html(html_path, _minimal_report_frames(user_table_success=True))

    html = html_path.read_text(encoding="utf-8")

    assert 'id="toggle-zero-user-roles"' in html
    assert 'data-zero-user-role-count="1"' in html
    assert "Show zero-user roles" in html
    assert "Hide zero-user roles" in html
    assert 'data-role="empty-role" data-zero-user-role="true" hidden' in html
    assert 'class="role-tab zero-user-role"' in html
    assert "syncZeroUserRoles" in html
    assert "document.querySelectorAll('.zero-user-role')" in html


def test_html_does_not_hide_zero_user_roles_when_user_table_failed(tmp_path):
    html_path = tmp_path / "report.html"
    _write_html(html_path, _minimal_report_frames(user_table_success=False))

    html = html_path.read_text(encoding="utf-8")

    assert 'id="toggle-zero-user-roles"' not in html
    assert 'data-zero-user-role="true"' not in html
    assert "Role user counts were not available from show user-table" in html


def test_html_does_not_hide_roles_when_all_roles_have_zero_users(tmp_path):
    html_path = tmp_path / "report.html"
    _write_html(html_path, _minimal_report_frames(user_table_success=True, active_user_count=0))

    html = html_path.read_text(encoding="utf-8")

    assert 'id="toggle-zero-user-roles"' not in html
    assert 'data-zero-user-role="true"' not in html
    assert "All Roles have 0 observed users" in html


def _minimal_report_frames(*, user_table_success: bool, active_user_count: int = 2) -> dict[str, pd.DataFrame]:
    role_network_rows = [
        {"role": "active-role", "observed_user_count": active_user_count},
        {"role": "empty-role", "observed_user_count": 0},
    ]
    acl_rows = [
        _acl_row("active-role", "active-role-acl", 1),
        _acl_row("empty-role", "empty-role-acl", 1),
    ]
    return {
        "Overview": pd.DataFrame(
            [{"controller": "sample", "ssid_count": 1, "role_count": 2, "alias_count": 0, "unresolved_count": 0}]
        ),
        "SSID_Role_Map": pd.DataFrame(),
        "Role_Network_Context": pd.DataFrame(role_network_rows),
        "Local_Role_Networks": pd.DataFrame(),
        "Role_ACL_Detail": pd.DataFrame(acl_rows),
        "Alias_Detail": pd.DataFrame(),
        "Unresolved": pd.DataFrame(),
        "Raw_Commands": pd.DataFrame(
            [{"controller": "sample", "command_id": "user_table", "command": "show user-table", "success": user_table_success}]
        ),
    }


def _acl_row(role: str, acl: str, sequence: int) -> dict[str, object]:
    return {
        "controller": "sample",
        "role": role,
        "acl": acl,
        "sequence": sequence,
        "action": "permit",
        "source": "user",
        "destination": "any",
        "source_interpretation": "User IP (current client assigned to this role)",
        "destination_interpretation": "Any (0.0.0.0/0)",
        "service": "any",
        "role_user_network": "",
        "access_summary": "",
        "access_flags": "",
        "raw_rule": "",
    }
