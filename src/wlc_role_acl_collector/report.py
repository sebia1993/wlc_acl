"""Build Excel and HTML reports from parsed controller data.

This module is also where privacy-sensitive export decisions are enforced, such
as not embedding local Role network mappings unless explicit export is enabled.
"""

from __future__ import annotations

import json
from datetime import datetime
from functools import lru_cache
from html import escape
from importlib.resources import files as resource_files
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .acl_evaluator import access_rule_id, build_access_check_data
from .aos8_parser import parse_controller_config
from .models import CollectionResult, ParsedController, RoleNetworkDefinition


def timestamp_slug() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def create_run_dir(output_root: Path, *, label: str = "") -> Path:
    output_root.mkdir(parents=True, exist_ok=True)
    label_suffix = f"_{_safe_file_name(label)}" if label else ""
    base_name = f"{timestamp_slug()}{label_suffix}"
    for counter in range(1000):
        name = base_name if counter == 0 else f"{base_name}_{counter:03d}"
        candidate = output_root / name
        try:
            candidate.mkdir()
            return candidate
        except FileExistsError:
            continue
    raise RuntimeError(f"Unable to create a unique run directory under {output_root}")


def write_raw_result(result: CollectionResult, raw_dir: Path) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / f"{_safe_file_name(result.controller.name)}.txt"
    lines = [
        "=== Controller ===",
        f"name: {result.controller.name}",
        f"host: {result.controller.host}",
        "",
    ]
    for command in result.commands:
        lines.extend(
            [
                f"=== Command: {command.command_id} | {command.command} ===",
                f"success: {command.success}",
                f"error: {command.error}",
                "--- output ---",
                _raw_output_for_storage(command),
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")
    result.raw_file = path
    return path


def _raw_output_for_storage(command) -> str:
    if command.command_id == "user_table":
        # user-table에는 사용자 IP/MAC 같은 민감정보가 섞일 수 있습니다.
        # 보고서에 필요한 건 사용자 수/대역 추정뿐이라 원문은 저장하지 않습니다.
        line_count = len([line for line in command.output.splitlines() if line.strip()])
        return "\n".join(
            [
                "[show user-table output redacted]",
                "Original output is used in memory only to calculate Role user count and observed network summary.",
                f"Non-empty line count: {line_count}",
            ]
        )
    return command.output or ""


def build_parsed_controllers(results: list[CollectionResult]) -> list[ParsedController]:
    parsed_items: list[ParsedController] = []
    for result in results:
        # collector.py는 명령 결과를 평평한 리스트로 모읍니다.
        # parser.py가 쓰기 쉽도록 Role별 rights 출력과 Alias별 netdestination 출력을 다시 묶습니다.
        rights_outputs = {}
        netdestination_outputs = {}
        for command in result.commands:
            if command.success and command.command_id.startswith("rights::"):
                rights_outputs[command.command_id.split("::", 1)[1]] = command.output
            if command.success and command.command_id.startswith("netdestination::"):
                netdestination_outputs[command.command_id.split("::", 1)[1]] = command.output
        parsed_items.append(
            parse_controller_config(
                controller=result.controller,
                config_text=result.command_output("configuration_effective"),
                rights_outputs=rights_outputs,
                netdestination_outputs=netdestination_outputs,
                ip_interface_brief_output=result.command_output("ip_interface_brief"),
                user_table_output=result.command_output("user_table"),
            )
        )
    return parsed_items


def write_reports(
    *,
    parsed_controllers: list[ParsedController],
    collection_results: list[CollectionResult],
    output_dir: Path,
    local_role_networks: list[RoleNetworkDefinition] | None = None,
    export_local_role_networks: bool = False,
    access_history_enabled: bool = False,
) -> dict[str, Path]:
    # 외부로 전달되는 최종 산출물은 여기서 만들어집니다.
    # local Role network는 사용자가 명시적으로 export를 켠 경우에만 보고서에 포함됩니다.
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "ssid_role_acl_report.xlsx"
    html_path = output_dir / "ssid_role_acl_report.html"

    local_role_networks = local_role_networks or []
    # pandas DataFrame은 Excel과 HTML 양쪽에서 재사용하는 중간 표 형식입니다.
    # 새 컬럼을 추가할 때는 이 frames 구조와 HTML/Excel 출력을 함께 확인해야 합니다.
    frames = _build_frames(
        parsed_controllers,
        collection_results,
        local_role_networks,
        export_local_role_networks=export_local_role_networks,
    )
    _write_excel(workbook_path, frames)
    _write_html(
        html_path,
        frames,
        local_role_networks_enabled=bool(local_role_networks) and export_local_role_networks,
        access_history_enabled=access_history_enabled,
    )
    return {"xlsx": workbook_path, "html": html_path}


def _build_frames(
    parsed_controllers: list[ParsedController],
    collection_results: list[CollectionResult],
    local_role_networks: list[RoleNetworkDefinition] | None = None,
    *,
    export_local_role_networks: bool = False,
) -> dict[str, pd.DataFrame]:
    local_role_networks = local_role_networks or []
    local_mapping_enabled = bool(local_role_networks) and export_local_role_networks
    local_lookup = _group_local_role_networks(local_role_networks) if local_mapping_enabled else {}
    ssid_rows: list[dict[str, Any]] = []
    role_network_rows: list[dict[str, Any]] = []
    local_network_rows: list[dict[str, Any]] = []
    acl_rows: list[dict[str, Any]] = []
    alias_rows: list[dict[str, Any]] = []
    unresolved_rows: list[dict[str, Any]] = []

    for parsed in parsed_controllers:
        local_status_by_role = _local_network_status_by_role(parsed, local_lookup, local_mapping_enabled)
        local_network_rows.extend(
            _local_network_rows(parsed, local_lookup, local_status_by_role, local_mapping_enabled)
        )
        for context in parsed.role_network_contexts:
            role_network_rows.append(
                {
                    "controller": context.controller,
                    "role": context.role,
                    "effective_vlan": context.effective_vlan,
                    "role_user_network": context.role_user_network,
                    "network_evidence": context.network_evidence,
                    "network_confidence": context.network_confidence,
                    "assignment_source": context.assignment_source,
                    "configured_vlan": context.configured_vlan,
                    "configured_subnet": context.configured_subnet,
                    "ssids": ", ".join(context.ssids),
                    "ap_groups": ", ".join(context.ap_groups),
                    "observed_user_count": context.observed_user_count,
                    "observed_vlans": ", ".join(context.observed_vlans),
                    "observed_networks": ", ".join(context.observed_networks),
                    "notes": context.notes,
                }
            )
        for mapping in parsed.ssid_role_mappings:
            local_status = local_status_by_role.get(mapping.role, _empty_local_network_status())
            ssid_row = {
                "controller": mapping.controller,
                "ssid": mapping.ssid,
                "ap_group": mapping.ap_group,
                "virtual_ap": mapping.virtual_ap,
                "ssid_profile": mapping.ssid_profile,
                "aaa_profile": mapping.aaa_profile,
                "role_type": mapping.role_type,
                "role": mapping.role,
                "vlan": mapping.vlan,
                "effective_vlan": mapping.effective_vlan,
                "role_user_network": mapping.role_user_network,
                "network_evidence": mapping.network_evidence,
                "network_confidence": mapping.network_confidence,
                "assignment_source": mapping.assignment_source,
                "configured_vlan": mapping.configured_vlan,
                "configured_subnet": mapping.configured_subnet,
                "observed_user_count": mapping.observed_user_count,
                "forward_mode": mapping.forward_mode,
                "access_summary": mapping.access_summary,
                "dynamic_role_possible": mapping.dynamic_role_possible,
                "dynamic_role_reason": mapping.dynamic_role_reason,
            }
            if local_mapping_enabled:
                ssid_row.update(
                    {
                        "local_role_networks": local_status["networks"],
                        "local_network_status": local_status["status"],
                        "local_network_notes": local_status["notes"],
                    }
                )
            ssid_rows.append(ssid_row)
        for policy in parsed.role_policies.values():
            local_status = local_status_by_role.get(policy.role, _empty_local_network_status())
            if not policy.rules:
                acl_row = {
                    "controller": policy.controller,
                    "role": policy.role,
                    "acl": ", ".join(policy.acl_names),
                    "sequence": "",
                    "action": "",
                    "source": "",
                    "destination": "",
                    "source_interpretation": "",
                    "destination_interpretation": "",
                    "service": "",
                    "role_user_network": _role_network_summary(parsed, policy.role),
                    "access_summary": policy.access_summary,
                    "access_flags": ", ".join(policy.access_flags),
                    "raw_rule": "",
                }
                if local_mapping_enabled:
                    acl_row.update(
                        {
                            "local_role_networks": local_status["networks"],
                            "local_network_status": local_status["status"],
                            "local_network_notes": local_status["notes"],
                        }
                    )
                acl_rows.append(acl_row)
            for rule in policy.rules:
                acl_row = {
                    "controller": policy.controller,
                    "role": policy.role,
                    "acl": rule.acl,
                    "sequence": rule.sequence,
                    "action": rule.action,
                    "source": rule.source,
                    "destination": rule.destination,
                    "source_interpretation": _acl_field_interpretation(
                        rule.source,
                        local_networks=local_status["networks"],
                        local_mapping_enabled=local_mapping_enabled,
                    ),
                    "destination_interpretation": _acl_field_interpretation(
                        rule.destination,
                        local_networks=local_status["networks"],
                        local_mapping_enabled=local_mapping_enabled,
                    ),
                    "service": rule.service,
                    "role_user_network": _role_network_summary(parsed, policy.role),
                    "access_summary": policy.access_summary,
                    "access_flags": ", ".join(policy.access_flags),
                    "raw_rule": rule.raw,
                }
                if local_mapping_enabled:
                    acl_row.update(
                        {
                            "local_role_networks": local_status["networks"],
                            "local_network_status": local_status["status"],
                            "local_network_notes": local_status["notes"],
                        }
                    )
                acl_rows.append(acl_row)
        for alias, entries in sorted(parsed.netdestination_aliases.items()):
            if not entries:
                alias_rows.append(
                    {
                        "controller": parsed.controller.name,
                        "alias": alias,
                        "sequence": "",
                        "entry_type": "",
                        "value": "",
                        "raw": "",
                    }
                )
            for entry in entries:
                alias_rows.append(
                    {
                        "controller": entry.controller,
                        "alias": entry.alias,
                        "sequence": entry.sequence,
                        "entry_type": entry.entry_type,
                        "value": entry.value,
                        "raw": entry.raw,
                    }
                )
        unresolved_rows.extend(parsed.unresolved)

    raw_rows: list[dict[str, Any]] = []
    for result in collection_results:
        raw_rows.extend(result.command_status_rows())

    overview_rows = []
    for parsed in parsed_controllers:
        controller = parsed.controller.name
        overview_rows.append(
            {
                "controller": controller,
                "ssid_count": len({row["ssid"] for row in ssid_rows if row["controller"] == controller}),
                "role_count": len(parsed.role_policies),
                "acl_rule_count": sum(
                    len(policy.rules) for policy in parsed.role_policies.values()
                ),
                "alias_count": len(parsed.netdestination_aliases),
                "unresolved_count": len(parsed.unresolved),
            }
        )

    frames = {
        "Overview": pd.DataFrame(overview_rows),
        "SSID_Role_Map": pd.DataFrame(ssid_rows),
        "Role_Network_Context": pd.DataFrame(role_network_rows),
        "Role_ACL_Detail": pd.DataFrame(acl_rows),
        "Alias_Detail": pd.DataFrame(alias_rows),
        "Unresolved": pd.DataFrame(unresolved_rows),
        "Raw_Commands": pd.DataFrame(raw_rows),
    }
    if local_mapping_enabled:
        frames["Local_Role_Networks"] = pd.DataFrame(
            local_network_rows,
            columns=[
                "controller",
                "role",
                "local_role_network",
                "subnet_mask",
                "source_file",
                "source_row",
                "status",
                "notes",
                "wlc_configured_subnets",
            ],
        )
    return frames


def _group_local_role_networks(
    definitions: list[RoleNetworkDefinition],
) -> dict[str, list[RoleNetworkDefinition]]:
    grouped: dict[str, list[RoleNetworkDefinition]] = {}
    for definition in definitions:
        grouped.setdefault(definition.role.casefold(), []).append(definition)
    return grouped


def _local_network_status_by_role(
    parsed: ParsedController,
    local_lookup: dict[str, list[RoleNetworkDefinition]],
    local_mapping_enabled: bool,
) -> dict[str, dict[str, str]]:
    if not local_mapping_enabled:
        return {}

    collected_roles = (
        set(parsed.role_policies)
        | set(parsed.user_role_observations)
        | {mapping.role for mapping in parsed.ssid_role_mappings if mapping.role}
    )
    collected_role_keys = {role.casefold() for role in collected_roles}
    local_roles = {
        definitions[0].role
        for definitions in local_lookup.values()
        if definitions and definitions[0].role.casefold() not in collected_role_keys
    }
    statuses: dict[str, dict[str, str]] = {}

    for role in sorted(collected_roles | local_roles, key=str.casefold):
        definitions = local_lookup.get(role.casefold(), [])
        local_networks = [definition.network for definition in definitions]
        wlc_networks = _wlc_configured_subnets_for_role(parsed, role)

        if role.casefold() not in collected_role_keys:
            status = "Role not collected"
            notes = "Role exists in the local Excel file, but it was not collected from this WLC."
        elif not definitions:
            status = "Local mapping missing"
            notes = "Role was collected from WLC but was not found in the local Role network Excel."
        elif wlc_networks and set(local_networks) != set(wlc_networks):
            status = "Mismatch"
            notes = f"Local: {', '.join(local_networks)} / WLC collected: {', '.join(wlc_networks)}"
        elif wlc_networks:
            status = "Matched"
            notes = "Local Role network matches WLC configured subnet evidence."
        else:
            status = "Local mapping loaded"
            notes = "No reliable WLC subnet was available for comparison."

        statuses[role] = {
            "networks": ", ".join(local_networks),
            "status": status,
            "notes": notes,
            "wlc_configured_subnets": ", ".join(wlc_networks),
        }
    return statuses


def _local_network_rows(
    parsed: ParsedController,
    local_lookup: dict[str, list[RoleNetworkDefinition]],
    status_by_role: dict[str, dict[str, str]],
    local_mapping_enabled: bool,
) -> list[dict[str, Any]]:
    if not local_mapping_enabled:
        return []

    rows: list[dict[str, Any]] = []
    for role, status in sorted(status_by_role.items(), key=lambda item: item[0].casefold()):
        definitions = local_lookup.get(role.casefold(), [])
        if not definitions:
            rows.append(
                {
                    "controller": parsed.controller.name,
                    "role": role,
                    "local_role_network": "",
                    "subnet_mask": "",
                    "source_file": "",
                    "source_row": "",
                    "status": status["status"],
                    "notes": status["notes"],
                    "wlc_configured_subnets": status["wlc_configured_subnets"],
                }
            )
            continue
        for definition in definitions:
            rows.append(
                {
                    "controller": parsed.controller.name,
                    "role": role,
                    "local_role_network": definition.network,
                    "subnet_mask": definition.subnet_mask,
                    "source_file": Path(definition.source_file).name,
                    "source_row": definition.source_row,
                    "status": status["status"],
                    "notes": status["notes"],
                    "wlc_configured_subnets": status["wlc_configured_subnets"],
                }
            )
    return rows


def _empty_local_network_status() -> dict[str, str]:
    return {
        "networks": "",
        "status": "",
        "notes": "",
        "wlc_configured_subnets": "",
    }


def _wlc_configured_subnets_for_role(parsed: ParsedController, role: str) -> list[str]:
    networks = [
        context.configured_subnet
        for context in parsed.role_network_contexts
        if context.role.casefold() == role.casefold()
        and context.configured_subnet
        and context.configured_subnet.casefold() != "unknown"
    ]
    return list(dict.fromkeys(networks))


def _role_network_summary(parsed: ParsedController, role: str) -> str:
    networks = [
        context.role_user_network
        for context in parsed.role_network_contexts
        if context.role == role and context.role_user_network
    ]
    return ", ".join(dict.fromkeys(networks)) or "Unknown"


def _acl_field_interpretation(
    value: str,
    *,
    local_networks: str = "",
    local_mapping_enabled: bool = False,
) -> str:
    normalized = value.strip().lower()
    if normalized == "user":
        if local_networks:
            return f"User IP (current client assigned to this role). Local Role Network: {local_networks}"
        if local_mapping_enabled:
            return "User IP (current client assigned to this role). Local Role Network: mapping missing."
        return "User IP (current client assigned to this role)"
    if normalized == "any":
        return "Any (0.0.0.0/0)"
    return ""


def _write_excel(path: Path, frames: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column_cells in worksheet.columns:
                max_length = 8
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 80))
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2


def _write_html(
    path: Path,
    frames: dict[str, pd.DataFrame],
    *,
    local_role_networks_enabled: bool = False,
    access_history_enabled: bool = False,
) -> None:
    overview = frames["Overview"].to_dict(orient="records")
    role_network_rows = frames["Role_Network_Context"].to_dict(orient="records")
    local_network_rows = frames.get("Local_Role_Networks", pd.DataFrame()).to_dict(orient="records")
    acl_rows = frames["Role_ACL_Detail"].to_dict(orient="records")
    alias_rows = frames["Alias_Detail"].to_dict(orient="records")
    alias_lookup = _group_by_alias(alias_rows)
    local_network_lookup = _group_local_network_rows(local_network_rows)
    role_user_counts = _role_observed_user_counts(role_network_rows)
    raw_command_rows = frames["Raw_Commands"].to_dict(orient="records")
    user_table_counts_reliable = _user_table_counts_reliable(raw_command_rows)
    unresolved_count = len(frames["Unresolved"])
    internal_role_network_banner = _internal_role_network_banner_html(
        local_network_rows,
        enabled=local_role_networks_enabled,
    )

    cards = "\n".join(
        f"""
        <section class="metric">
          <span class="metric-label">{escape(str(row.get('controller', '')))}</span>
          <strong class="metric-value">{escape(str(row.get('ssid_count', 0)))}</strong>
          <small class="metric-detail">SSID / Role {escape(str(row.get('role_count', 0)))} / Alias {escape(str(row.get('alias_count', 0)))} / Unresolved {escape(str(row.get('unresolved_count', 0)))}</small>
        </section>
        """
        for row in overview
    )

    acl_by_role: dict[str, list[dict[str, Any]]] = {}
    for row in acl_rows:
        acl_by_role.setdefault(str(row.get("role", "")), []).append(row)

    # HTML에서는 Role을 표 행으로 길게 나열하지 않고 탭 버튼으로 보여줍니다.
    # 운영자가 자주 볼 가능성이 높은 "관측 사용자 수가 많은 Role"을 앞에 배치합니다.
    role_items = [
        {
            "role": role,
            "rows": rows,
            "other_acl_count": _other_acl_row_count(role, rows),
            "user_count": role_user_counts.get(role, 0),
            "panel_id": f"role-panel-{index}",
        }
        for index, (role, rows) in enumerate(
            sorted(
                acl_by_role.items(),
                key=lambda item: (-role_user_counts.get(item[0], 0), item[0].lower()),
            ),
            start=1,
        )
    ]
    executive_summary = _executive_summary_html(
        frames,
        role_items,
        local_network_rows,
        unresolved_count=unresolved_count,
    )
    zero_user_hiding_enabled = (
        user_table_counts_reliable
        and any(_int_value(item["user_count"]) > 0 for item in role_items)
        and any(_int_value(item["user_count"]) == 0 for item in role_items)
    )
    # user-table이 신뢰 가능할 때만 사용자 0명 Role을 기본 숨김 처리합니다.
    # 데이터가 없을 때 전부 숨겨져 빈 보고서가 되는 상황을 피하기 위한 조건입니다.
    for item in role_items:
        item["zero_user_hidden"] = zero_user_hiding_enabled and _int_value(item["user_count"]) == 0
    zero_user_hidden_count = sum(1 for item in role_items if item["zero_user_hidden"])
    selected_panel_id = next(
        (str(item["panel_id"]) for item in role_items if not item["zero_user_hidden"]),
        str(role_items[0]["panel_id"]) if role_items else "",
    )
    zero_user_role_controls = _zero_user_role_controls_html(
        enabled=zero_user_hiding_enabled,
        hidden_count=zero_user_hidden_count,
        user_table_counts_reliable=user_table_counts_reliable,
        role_items=role_items,
    )
    access_check_data = build_access_check_data(
        role_items,
        alias_rows,
        local_network_rows,
        include_local_networks=local_role_networks_enabled,
    )
    # Access Check는 브라우저 안에서 동작하므로 필요한 ACL/Alias 데이터를 JSON으로 HTML에 심습니다.
    # local Role network는 보안 정책상 export 옵션이 켜진 경우에만 포함됩니다.
    access_check_json = _json_for_html(access_check_data)
    access_check_controls = _access_check_controls_html(
        access_check_data,
        history_enabled=access_history_enabled,
        local_role_networks_enabled=local_role_networks_enabled,
    )
    access_check_css = _access_check_css()
    access_check_script = _access_check_script(history_enabled=access_history_enabled)
    role_image_export_script = _role_image_export_script()
    html2canvas_source = _html2canvas_source()
    role_buttons = "\n".join(
        f"""
        <button class="role-tab{_zero_user_role_class(bool(item['zero_user_hidden']))}" type="button" role="tab" data-role="{escape(str(item['role']))}" data-panel-id="{escape(str(item['panel_id']))}"
          aria-controls="{escape(str(item['panel_id']))}" aria-selected="{'true' if str(item['panel_id']) == selected_panel_id else 'false'}"{_zero_user_role_attrs(bool(item['zero_user_hidden']))}{_hidden_attr(bool(item['zero_user_hidden']))}>
          <span class="role-tab-name">{escape(str(item['role']))}</span>
          <span class="role-tab-meta">{len(item['rows'])} rules / {item['user_count']} users</span>
        </button>
        """
        for index, item in enumerate(role_items, start=1)
    )
    acl_sections = "\n".join(
        f"""
        <section class="acl-section role-panel{_zero_user_role_class(bool(item['zero_user_hidden']))}" id="{escape(str(item['panel_id']))}" data-role="{escape(str(item['role']))}"{_zero_user_role_attrs(bool(item['zero_user_hidden']))}{_hidden_attr(bool(item['zero_user_hidden']) or str(item['panel_id']) != selected_panel_id)}>
          <div class="acl-section-header">
            <h3>{escape(str(item['role']))}</h3>
            <span>{len(item['rows'])} rules / {item['user_count']} observed users{_other_acl_meta_text(int(item['other_acl_count']))}</span>
          </div>
          {_role_description_html(str(item['role']))}
          {_local_role_network_html(local_network_lookup.get(str(item['role']), []), local_role_networks_enabled)}
          {_other_acl_toggle_html(str(item['panel_id']), int(item['other_acl_count']))}
          <table>
            <thead><tr><th>ACL</th><th>#</th><th>Action</th><th>Source</th><th>Destination</th><th>Service</th><th class="raw-column">Raw</th><th>Comment</th></tr></thead>
            <tbody>
              {_acl_rows_html(str(item['role']), item['rows'], alias_lookup)}
            </tbody>
          </table>
        </section>
        """
        for index, item in enumerate(role_items, start=1)
    )

    html = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WLC SSID Role ACL Report</title>
  <style>
    :root {{
      --bg: #eef2f6;
      --panel: #ffffff;
      --panel-subtle: #f8fafc;
      --text: #172033;
      --muted: #667085;
      --line: #d6dde7;
      --line-strong: #b8c3d2;
      --accent: #0f6cbd;
      --accent-deep: #174a7c;
      --accent-soft: #eaf3ff;
      --success: #067647;
      --success-soft: #ecfdf3;
      --danger: #b42318;
      --danger-soft: #fef3f2;
      --warning: #b54708;
      --warning-soft: #fffaeb;
      --info: #0b7a75;
      --shadow: 0 12px 28px rgba(15, 23, 42, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Malgun Gothic", Arial, sans-serif;
      color: var(--text);
      background:
        linear-gradient(180deg, #f8fafc 0, var(--bg) 280px),
        var(--bg);
    }}
    header {{
      background: #ffffff;
      border-bottom: 1px solid var(--line);
      padding: 18px 28px;
    }}
    .report-header-inner {{
      align-items: center;
      display: flex;
      gap: 18px;
      justify-content: space-between;
      margin: 0 auto;
      max-width: 1480px;
    }}
    .report-kicker {{
      color: var(--accent-deep);
      font-size: 12px;
      font-weight: 800;
      letter-spacing: 0;
      margin-bottom: 5px;
      text-transform: uppercase;
    }}
    .report-summary-pill {{
      align-items: center;
      background: var(--accent-soft);
      border: 1px solid #b9dcff;
      border-radius: 8px;
      color: var(--accent-deep);
      display: inline-flex;
      gap: 10px;
      padding: 10px 12px;
    }}
    .report-summary-pill span {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }}
    .report-summary-pill strong {{ font-size: 22px; }}
    h1 {{ margin: 0; font-size: 25px; letter-spacing: 0; }}
    h2 {{ font-size: 18px; margin: 22px 0 10px; }}
    main {{
      margin: 0 auto;
      max-width: 1480px;
      padding: 22px 28px 42px;
    }}
    .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px; margin-bottom: 16px; }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
      padding: 14px 15px;
      position: relative;
    }}
    .metric::before {{
      background: var(--accent);
      border-radius: 8px 0 0 8px;
      content: "";
      inset: 0 auto 0 0;
      position: absolute;
      width: 4px;
    }}
    .metric-label, .metric-detail {{ color: var(--muted); display: block; }}
    .metric-label {{ font-size: 12px; font-weight: 700; }}
    .metric-value {{ font-size: 30px; display: block; margin: 4px 0; }}
    .metric-detail {{ font-size: 12px; }}
    .executive-summary {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      margin: 0 0 16px;
      padding: 16px;
    }}
    .executive-summary-header {{
      align-items: flex-start;
      display: flex;
      gap: 12px;
      justify-content: space-between;
      margin-bottom: 12px;
    }}
    .executive-summary h2 {{
      font-size: 18px;
      margin: 0;
    }}
    .executive-summary p {{
      color: var(--muted);
      margin: 4px 0 0;
    }}
    .attention-badge {{
      background: var(--warning-soft);
      border: 1px solid #fedf89;
      border-radius: 999px;
      color: var(--warning);
      flex: 0 0 auto;
      font-size: 12px;
      font-weight: 800;
      padding: 6px 10px;
    }}
    .summary-grid {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(210px, 1fr));
    }}
    .summary-item {{
      background: var(--panel-subtle);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
    }}
    .summary-item strong {{
      display: block;
      font-size: 14px;
      margin-bottom: 5px;
    }}
    .summary-item span,
    .summary-item li {{
      color: var(--muted);
      font-size: 12px;
      line-height: 1.5;
    }}
    .summary-item ul {{
      margin: 0;
      padding-left: 18px;
    }}
    .internal-network-banner {{
      background: #fff7ed;
      border: 1px solid #fed7aa;
      border-radius: 8px;
      color: #7c2d12;
      margin: 0 0 16px;
      padding: 13px 15px;
    }}
    .internal-network-banner strong {{
      display: block;
      font-size: 14px;
      margin-bottom: 5px;
    }}
    .internal-network-banner p {{
      margin: 0 0 8px;
    }}
    .internal-network-banner ul {{
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      list-style: none;
      margin: 0;
      padding: 0;
    }}
    .internal-network-banner li {{
      background: #ffffff;
      border: 1px solid #fed7aa;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      padding: 4px 8px;
    }}
    .report-actions {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 16px 0 12px;
    }}
    .report-action {{
      background: #0f6cbd;
      border: 0;
      border-radius: 6px;
      color: #ffffff;
      cursor: pointer;
      font-size: 13px;
      font-weight: 700;
      padding: 10px 12px;
      transition: background .15s ease, box-shadow .15s ease, transform .15s ease;
    }}
    .report-action:hover {{
      background: #0b5aa1;
      box-shadow: 0 6px 14px rgba(15, 108, 189, 0.22);
      transform: translateY(-1px);
    }}
    .report-action:disabled {{
      box-shadow: none;
      cursor: wait;
      opacity: .65;
      transform: none;
    }}
    .report-action.secondary {{ background: #e8eef6; color: #15324b; }}
    .report-action.secondary:hover {{
      background: #dce6f2;
      box-shadow: 0 6px 14px rgba(21, 50, 75, 0.10);
    }}
    .role-image-status {{
      align-self: center;
      color: var(--muted);
      font-size: 12px;
      min-height: 18px;
    }}
    .role-image-status[data-state="success"] {{ color: var(--success); }}
    .role-image-status[data-state="error"] {{ color: var(--danger); }}
    .role-tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 10px 0 14px;
    }}
    .role-tab {{
      align-items: flex-start;
      background: #ffffff;
      border: 1px solid var(--line);
      border-radius: 8px;
      color: var(--text);
      cursor: pointer;
      display: inline-flex;
      flex-direction: column;
      gap: 3px;
      min-width: 150px;
      padding: 10px 12px;
      position: relative;
      text-align: left;
      transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease;
    }}
    .role-tab:hover {{
      border-color: #9fc5f1;
      box-shadow: 0 6px 14px rgba(15, 23, 42, 0.08);
      transform: translateY(-1px);
    }}
    .role-tab[aria-selected="true"] {{
      background: #eef6ff;
      border-color: #5aa9f4;
      box-shadow: 0 8px 18px rgba(15, 108, 189, 0.14);
      color: #175cd3;
      font-weight: 600;
    }}
    .role-tab[aria-selected="true"]::after {{
      background: var(--accent);
      border-radius: 999px;
      bottom: 7px;
      content: "";
      height: 5px;
      position: absolute;
      right: 8px;
      width: 5px;
    }}
    .role-tab-name {{ font-size: 13px; line-height: 1.25; }}
    .role-tab-meta {{ color: var(--muted); font-size: 11px; font-weight: 400; }}
    .zero-user-role[hidden] {{ display: none; }}
    .zero-user-role-controls {{
      align-items: center;
      display: flex;
      gap: 8px;
      margin: 6px 0 10px;
    }}
    .zero-user-role-notice {{
      background: #fff7ed;
      border: 1px solid #fed7aa;
      border-radius: 6px;
      color: #9a3412;
      font-size: 12px;
      margin: 6px 0 10px;
      padding: 8px 10px;
    }}
    .role-report-description {{
      background: #f8fbff;
      border-top: 1px solid var(--line);
      padding: 12px 14px;
    }}
    .role-description-header {{
      align-items: center;
      display: flex;
      gap: 10px;
      justify-content: space-between;
      margin-bottom: 7px;
    }}
    .role-description-header label {{
      color: #15324b;
      font-size: 13px;
      font-weight: 700;
    }}
    .role-description-input {{
      border: 1px solid var(--line);
      border-radius: 6px;
      font-family: inherit;
      font-size: 13px;
      min-height: 88px;
      padding: 9px 10px;
      resize: vertical;
      width: 100%;
    }}
    .role-description-status {{
      color: var(--muted);
      font-size: 11px;
    }}
    .role-description-status[data-state="saved"],
    .role-description-status[data-state="restored"] {{
      color: #05603a;
    }}
    .role-description-status[data-state="unavailable"] {{
      color: #b42318;
    }}
    .role-description-print {{
      display: none;
      min-height: 22px;
      white-space: pre-wrap;
    }}
    .local-network {{
      background: var(--panel-subtle);
      border-top: 1px solid var(--line);
      padding: 10px 14px 12px;
    }}
    .local-network-title {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      margin-bottom: 7px;
      text-transform: uppercase;
    }}
    .local-subnet-pill {{
      align-items: center;
      background: #ffffff;
      border: 1px solid var(--line);
      border-radius: 999px;
      display: inline-flex;
      gap: 6px;
      margin: 3px 6px 3px 0;
      padding: 4px 8px;
    }}
    .local-network-status {{
      border-radius: 999px;
      color: #ffffff;
      display: inline-block;
      font-size: 11px;
      font-weight: 700;
      margin: 3px 6px 3px 0;
      padding: 4px 8px;
    }}
    .local-network-status.status-matched {{ background: #05603a; }}
    .local-network-status.status-missing {{ background: #b42318; }}
    .local-network-status.status-mismatch {{ background: #c2410c; }}
    .local-network-status.status-loaded {{ background: #0f6cbd; }}
    .local-network-status.status-not-collected {{ background: #667085; }}
    .local-network-notes {{ color: var(--muted); font-size: 12px; margin-top: 5px; }}
    .acl-filter-actions {{
      background: var(--panel-subtle);
      border-top: 1px solid var(--line);
      padding: 9px 14px;
    }}
    .acl-filter-button {{
      background: #e8eef6;
      border: 1px solid #d4deea;
      border-radius: 6px;
      color: #15324b;
      cursor: pointer;
      font-size: 12px;
      font-weight: 700;
      padding: 8px 10px;
    }}
    .other-acl-rule[hidden] {{ display: none; }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; background: var(--panel); }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 9px 10px; text-align: left; vertical-align: top; font-size: 13px; }}
    th {{
      background: #213b59;
      color: #fff;
      font-size: 12px;
      font-weight: 800;
      position: sticky;
      top: 0;
      z-index: 2;
    }}
    tbody tr:nth-child(even):not(.alias-detail-row) {{ background: #fbfcfe; }}
    tbody tr:hover:not(.alias-detail-row) {{ background: #f0f6ff; }}
    .raw-column {{ display: none; }}
    body.raw-visible .raw-column {{ display: table-cell; }}
    .alias-prefix {{
      color: var(--muted);
      margin-right: 4px;
    }}
    .alias-link {{
      border: 1px solid #a6c8ff;
      background: #eef6ff;
      color: #0f4c81;
      border-radius: 999px;
      cursor: pointer;
      font: inherit;
      padding: 2px 8px;
    }}
    .alias-link:hover {{ background: #dbeafe; }}
    .alias-detail-row[hidden] {{ display: none; }}
    .alias-detail {{
      background: #f8fafc;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px 12px;
    }}
    .alias-detail-title {{
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 8px;
      text-transform: uppercase;
    }}
    .alias-chip {{
      align-items: center;
      background: #ffffff;
      border: 1px solid var(--line);
      border-radius: 999px;
      display: inline-flex;
      gap: 6px;
      margin: 3px 6px 3px 0;
      padding: 4px 8px;
    }}
    .alias-type {{
      border-radius: 999px;
      color: #ffffff;
      font-size: 11px;
      font-weight: 700;
      letter-spacing: 0;
      padding: 2px 6px;
    }}
    .alias-type-host {{ background: #0f6cbd; }}
    .alias-type-network {{ background: #0b7a75; }}
    .alias-type-range {{ background: #c2410c; }}
    .alias-type-name {{ background: #6d28d9; }}
    .alias-type-raw {{ background: #667085; }}
    .rule-badge {{
      border-radius: 999px;
      display: inline-flex;
      font-size: 11px;
      font-weight: 800;
      line-height: 1;
      padding: 5px 8px;
      white-space: nowrap;
    }}
    .action-permit {{ background: var(--success-soft); color: var(--success); }}
    .action-deny {{ background: var(--danger-soft); color: var(--danger); }}
    .action-special {{ background: var(--warning-soft); color: var(--warning); }}
    .action-unknown {{ background: #eef2f6; color: #475467; }}
    .service-badge {{
      background: #f2f4f7;
      border: 1px solid #e4e7ec;
      color: #344054;
      max-width: 180px;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .acl-section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      margin-top: 14px;
      overflow: hidden;
    }}
    .acl-section-header {{
      align-items: center;
      background: linear-gradient(180deg, #ffffff, #f8fafc);
      border-bottom: 1px solid var(--line);
      display: flex;
      gap: 10px;
      justify-content: space-between;
      padding: 12px 14px;
    }}
    .acl-section-header h3 {{ font-size: 16px; margin: 0; }}
    .acl-section-header span {{ color: var(--muted); font-size: 12px; }}
    .role-image-export {{
      box-shadow: none;
      margin: 0;
      max-width: none;
      overflow: visible;
    }}
    .role-image-export .comment-export-text,
    .role-image-export .role-description-export-text {{
      min-height: 20px;
      overflow-wrap: anywhere;
      white-space: pre-wrap;
    }}
    .role-image-export .role-description-export-text {{
      color: var(--text);
      line-height: 1.55;
    }}
    .role-image-export .alias-link {{
      background: transparent;
      border: 0;
      color: inherit;
      cursor: default;
      padding: 0;
    }}
    .role-image-export .image-export-part {{
      color: var(--muted);
      font-size: 11px;
      font-weight: 700;
      margin-left: 8px;
    }}
    .raw {{ font-family: Consolas, monospace; white-space: pre-wrap; }}
    .comment-cell {{ min-width: 220px; width: 24%; }}
    .comment-input {{
      border: 1px solid var(--line);
      border-radius: 6px;
      font-family: inherit;
      font-size: 13px;
      min-height: 74px;
      padding: 8px 10px;
      resize: vertical;
      width: 100%;
    }}
    .comment-print {{
      display: none;
      white-space: pre-wrap;
    }}
    .comment-status {{
      color: var(--muted);
      font-size: 11px;
      margin-top: 4px;
    }}
    .comment-status[data-state="saved"],
    .comment-status[data-state="restored"] {{
      color: #05603a;
    }}
    .comment-status[data-state="unavailable"] {{
      color: #b42318;
    }}
    .notice {{ color: var(--muted); margin: 8px 0 0; }}
    {access_check_css}
    @media (max-width: 720px) {{
      header {{ padding: 16px 18px; }}
      main {{ padding: 18px 16px 34px; }}
      .report-header-inner {{
        align-items: flex-start;
        flex-direction: column;
      }}
      .report-summary-pill {{
        width: 100%;
        justify-content: space-between;
      }}
      .role-tab {{
        flex: 1 1 100%;
      }}
      .acl-section-header {{
        align-items: flex-start;
        flex-direction: column;
      }}
      .role-description-header {{
        align-items: flex-start;
        flex-direction: column;
      }}
      th, td {{
        padding: 8px;
      }}
    }}
    @media print {{
      body {{ background: #ffffff; }}
      header, main {{ padding: 14px 18px; }}
      .no-print {{ display: none !important; }}
      .metric,
      .executive-summary,
      .acl-section,
      .access-check {{
        box-shadow: none;
      }}
      .acl-section {{ break-inside: auto; }}
      th {{ position: static; }}
      .alias-link {{
        background: transparent;
        border: 0;
        color: inherit;
        padding: 0;
      }}
      .comment-input {{ display: none; }}
      .comment-print {{
        display: block;
        min-height: 20px;
        white-space: pre-wrap;
      }}
      .comment-status {{ display: none; }}
      .role-description-input {{ display: none; }}
      .role-description-print {{
        display: block;
        min-height: 22px;
        white-space: pre-wrap;
      }}
      .role-description-status {{ display: none; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="report-header-inner">
      <div>
        <div class="report-kicker">Wireless Role ACL Review</div>
        <h1>WLC SSID Role ACL Report</h1>
        <p class="notice">생성시각: {escape(datetime.now().isoformat(timespec='seconds'))} / Unresolved: {unresolved_count}</p>
      </div>
      <div class="report-summary-pill">
        <span>Controllers</span>
        <strong>{len(overview)}</strong>
      </div>
    </div>
  </header>
  <main>
    <div class="metrics">{cards}</div>
    {executive_summary}
    {internal_role_network_banner}
    <h2>Role ACL Detail</h2>
    <div class="report-actions no-print">
      <button id="save-commented-html" class="report-action" type="button">주석 포함 HTML 저장</button>
      <button id="save-role-png" class="report-action" type="button">선택 Role PNG 저장</button>
      <button id="print-pdf" class="report-action secondary" type="button">PDF 저장/인쇄</button>
      <button id="toggle-raw" class="report-action secondary" type="button" aria-pressed="false">Raw 보기</button>
      <span id="role-image-status" class="role-image-status" aria-live="polite"></span>
    </div>
    {zero_user_role_controls}
    <div class="role-tabs no-print" role="tablist" aria-label="Role ACL list">
      {role_buttons}
    </div>
    {acl_sections}
    {access_check_controls}
  </main>
  <script id="access-check-data" type="application/json">{access_check_json}</script>
  {_access_check_history_data_html(access_history_enabled)}
  <textarea id="acl-comments-data" hidden>{{}}</textarea>
  <textarea id="role-descriptions-data" hidden>{{}}</textarea>
  <script>{html2canvas_source}</script>
  <script>
    const roleTabs = Array.from(document.querySelectorAll('.role-tab'));
    const rolePanels = Array.from(document.querySelectorAll('.role-panel'));
    const rawToggleButton = document.querySelector('#toggle-raw');
    const zeroUserToggleButton = document.querySelector('#toggle-zero-user-roles');
    const otherAclToggles = Array.from(document.querySelectorAll('.toggle-other-acls'));
    let selectedRolePanelId = roleTabs.find((button) => button.getAttribute('aria-selected') === 'true' && !button.hidden)?.dataset.panelId || roleTabs.find((button) => !button.hidden)?.dataset.panelId || roleTabs[0]?.dataset.panelId || '';
    {access_check_script}
    {role_image_export_script}

    function syncRawToggleButton() {{
      if (!rawToggleButton) {{
        return;
      }}
      const rawVisible = document.body.classList.contains('raw-visible');
      rawToggleButton.textContent = rawVisible ? 'Raw 숨김' : 'Raw 보기';
      rawToggleButton.setAttribute('aria-pressed', String(rawVisible));
    }}

    if (rawToggleButton) {{
      rawToggleButton.addEventListener('click', () => {{
        document.body.classList.toggle('raw-visible');
        syncRawToggleButton();
      }});
      syncRawToggleButton();
    }}

    function zeroUserRolesVisible() {{
      return !zeroUserToggleButton || zeroUserToggleButton.getAttribute('aria-pressed') === 'true';
    }}

    function firstVisibleRolePanelId() {{
      return roleTabs.find((button) => !button.hidden)?.dataset.panelId || roleTabs[0]?.dataset.panelId || '';
    }}

    function rolePanelIsAvailable(panelId) {{
      const button = roleTabs.find((item) => item.dataset.panelId === panelId);
      return Boolean(button && !button.hidden);
    }}

    function syncZeroUserToggleButton() {{
      if (!zeroUserToggleButton) {{
        return;
      }}
      const count = zeroUserToggleButton.dataset.zeroUserRoleCount || '0';
      const visible = zeroUserRolesVisible();
      zeroUserToggleButton.textContent = visible ? `Hide zero-user roles (${{count}})` : `Show zero-user roles (${{count}})`;
      zeroUserToggleButton.setAttribute('aria-pressed', String(visible));
    }}

    function syncZeroUserRoles() {{
      const visible = zeroUserRolesVisible();
      for (const item of document.querySelectorAll('.zero-user-role')) {{
        item.hidden = !visible;
      }}
      if (!rolePanelIsAvailable(selectedRolePanelId)) {{
        selectedRolePanelId = firstVisibleRolePanelId();
      }}
      if (selectedRolePanelId) {{
        selectRolePanel(selectedRolePanelId);
      }}
      syncZeroUserToggleButton();
    }}

    if (zeroUserToggleButton) {{
      zeroUserToggleButton.addEventListener('click', () => {{
        zeroUserToggleButton.setAttribute('aria-pressed', String(!zeroUserRolesVisible()));
        syncZeroUserRoles();
      }});
    }}

    function syncOtherAclButton(button) {{
      const count = button.dataset.otherAclCount || '0';
      const visible = button.getAttribute('aria-pressed') === 'true';
      button.textContent = visible ? `Hide other ACLs (${{count}})` : `Show other ACLs (${{count}})`;
    }}

    function syncOtherAclRows(button) {{
      const panel = button.closest('.role-panel');
      if (!panel) {{
        return;
      }}
      const visible = button.getAttribute('aria-pressed') === 'true';
      for (const row of panel.querySelectorAll('.other-acl-rule')) {{
        row.hidden = !visible;
      }}
      if (!visible) {{
        for (const link of panel.querySelectorAll('.other-acl-rule .alias-link')) {{
          const detail = document.getElementById(link.dataset.detailId);
          if (detail) {{
            detail.hidden = true;
          }}
          link.setAttribute('aria-expanded', 'false');
        }}
      }}
      syncOtherAclButton(button);
    }}

    for (const button of otherAclToggles) {{
      button.addEventListener('click', () => {{
        const visible = button.getAttribute('aria-pressed') === 'true';
        button.setAttribute('aria-pressed', String(!visible));
        syncOtherAclRows(button);
      }});
      syncOtherAclRows(button);
    }}

    function selectRolePanel(panelId) {{
      selectedRolePanelId = rolePanelIsAvailable(panelId) ? panelId : firstVisibleRolePanelId();
      for (const panel of rolePanels) {{
        const available = !panel.classList.contains('zero-user-role') || zeroUserRolesVisible();
        panel.hidden = !available || panel.id !== selectedRolePanelId;
      }}
      for (const button of roleTabs) {{
        button.setAttribute('aria-selected', String(button.dataset.panelId === selectedRolePanelId));
      }}
      const selectedPanel = rolePanels.find((panel) => panel.id === selectedRolePanelId);
      if (selectedPanel && typeof syncAccessRoleSelection === 'function') {{
        syncAccessRoleSelection(selectedPanel.dataset.role || '');
      }}
    }}

    for (const button of roleTabs) {{
      button.addEventListener('click', () => {{
        selectRolePanel(button.dataset.panelId);
      }});
    }}
    syncZeroUserRoles();

    for (const button of document.querySelectorAll('.alias-link')) {{
      button.addEventListener('click', () => {{
        const detailId = button.dataset.detailId;
        const detail = document.getElementById(detailId);
        if (!detail) return;
        const isHidden = detail.hasAttribute('hidden');
        detail.toggleAttribute('hidden', !isHidden);
        button.setAttribute('aria-expanded', String(isHidden));
      }});
    }}
    const commentsDataElement = document.querySelector('#acl-comments-data');
    const commentInputs = Array.from(document.querySelectorAll('.comment-input'));
    const roleDescriptionsDataElement = document.querySelector('#role-descriptions-data');
    const roleDescriptionInputs = Array.from(document.querySelectorAll('.role-description-input'));
    const commentStorageKey = `wlc-role-acl-comments:${{location.pathname || document.title}}`;
    const roleDescriptionStorageKey = `wlc-role-report-descriptions:${{location.pathname || document.title}}`;
    const commentStorageAvailable = (() => {{
      try {{
        localStorage.setItem('wlc-role-acl-storage-test', '1');
        localStorage.removeItem('wlc-role-acl-storage-test');
        return true;
      }} catch (_error) {{
        return false;
      }}
    }})();
    let aclComments = {{}};
    let roleDescriptions = {{}};

    function readEmbeddedComments() {{
      try {{
        return JSON.parse(commentsDataElement.value || commentsDataElement.textContent || '{{}}');
      }} catch (_error) {{
        return {{}};
      }}
    }}

    function syncCommentsData() {{
      const text = JSON.stringify(aclComments, null, 2);
      commentsDataElement.value = text;
      commentsDataElement.textContent = text;
    }}

    function readStoredComments() {{
      if (!commentStorageAvailable) {{
        return {{}};
      }}
      try {{
        return JSON.parse(localStorage.getItem(commentStorageKey) || '{{}}');
      }} catch (_error) {{
        return {{}};
      }}
    }}

    function persistComments() {{
      if (!commentStorageAvailable) {{
        return;
      }}
      localStorage.setItem(commentStorageKey, JSON.stringify(aclComments));
    }}

    function readEmbeddedRoleDescriptions() {{
      try {{
        return JSON.parse(
          roleDescriptionsDataElement.value
          || roleDescriptionsDataElement.textContent
          || '{{}}'
        );
      }} catch (_error) {{
        return {{}};
      }}
    }}

    function syncRoleDescriptionsData() {{
      const text = JSON.stringify(roleDescriptions, null, 2);
      roleDescriptionsDataElement.value = text;
      roleDescriptionsDataElement.textContent = text;
    }}

    function readStoredRoleDescriptions() {{
      if (!commentStorageAvailable) {{
        return {{}};
      }}
      try {{
        return JSON.parse(localStorage.getItem(roleDescriptionStorageKey) || '{{}}');
      }} catch (_error) {{
        return {{}};
      }}
    }}

    function persistRoleDescriptions() {{
      if (!commentStorageAvailable) {{
        return;
      }}
      localStorage.setItem(roleDescriptionStorageKey, JSON.stringify(roleDescriptions));
    }}

    function setRoleDescriptionStatus(input, text, state) {{
      const status = document.getElementById(`${{input.dataset.descriptionId}}-status`);
      if (!status) {{
        return;
      }}
      status.textContent = text;
      status.dataset.state = state || '';
    }}

    function updateRoleDescriptionPrint(input) {{
      const printValue = document.getElementById(`${{input.dataset.descriptionId}}-print`);
      if (printValue) {{
        printValue.textContent = input.value || '입력된 설명이 없습니다.';
      }}
    }}

    function collectRoleDescriptions() {{
      roleDescriptions = {{}};
      for (const input of roleDescriptionInputs) {{
        if (input.value) {{
          roleDescriptions[input.dataset.role] = input.value;
        }}
        updateRoleDescriptionPrint(input);
      }}
      syncRoleDescriptionsData();
      persistRoleDescriptions();
    }}

    function syncRoleDescriptionDomValues() {{
      for (const input of roleDescriptionInputs) {{
        input.textContent = input.value;
      }}
    }}

    function setCommentStatus(input, text, state) {{
      const status = document.getElementById(`${{input.dataset.commentId}}-status`);
      if (!status) {{
        return;
      }}
      status.textContent = text;
      status.dataset.state = state || '';
    }}

    function updateCommentPrint(input) {{
      const printValue = document.getElementById(`${{input.dataset.commentId}}-print`);
      if (printValue) {{
        printValue.textContent = input.value || '';
      }}
    }}

    function collectComments() {{
      aclComments = {{}};
      for (const input of commentInputs) {{
        if (input.value) {{
          aclComments[input.dataset.commentId] = input.value;
        }}
        updateCommentPrint(input);
      }}
      syncCommentsData();
      persistComments();
    }}

    function syncTextareaDomValues() {{
      for (const input of commentInputs) {{
        input.textContent = input.value;
      }}
    }}

    function preparePrint() {{
      collectComments();
      collectRoleDescriptions();
      for (const panel of rolePanels) {{
        panel.hidden = false;
      }}
      for (const item of document.querySelectorAll('.zero-user-role')) {{
        item.hidden = false;
      }}
      for (const row of document.querySelectorAll('.other-acl-rule')) {{
        row.hidden = false;
      }}
      for (const detail of document.querySelectorAll('.alias-detail-row')) {{
        detail.hidden = false;
      }}
      for (const button of document.querySelectorAll('.alias-link')) {{
        button.setAttribute('aria-expanded', 'true');
      }}
    }}

    function restoreScreenView() {{
      syncZeroUserRoles();
      for (const button of otherAclToggles) {{
        syncOtherAclRows(button);
      }}
    }}

    function saveCommentedHtml() {{
      collectComments();
      collectRoleDescriptions();
      syncTextareaDomValues();
      syncRoleDescriptionDomValues();
      if (typeof syncAccessHistoryDomValues === 'function') {{
        syncAccessHistoryDomValues();
      }}
      const html = '<!doctype html>\\n' + document.documentElement.outerHTML;
      const blob = new Blob([html], {{ type: 'text/html;charset=utf-8' }});
      const stamp = new Date().toISOString().replace(/[:.]/g, '-');
      const link = document.createElement('a');
      link.href = URL.createObjectURL(blob);
      link.download = `ssid_role_acl_report_commented_${{stamp}}.html`;
      document.body.appendChild(link);
      link.click();
      URL.revokeObjectURL(link.href);
      link.remove();
    }}

    const embeddedComments = readEmbeddedComments();
    const storedComments = readStoredComments();
    const hasStoredComments = Object.keys(storedComments).length > 0;
    aclComments = Object.assign({{}}, embeddedComments, storedComments);
    for (const input of commentInputs) {{
      input.value = aclComments[input.dataset.commentId] || '';
      updateCommentPrint(input);
      if (!commentStorageAvailable) {{
        setCommentStatus(input, '브라우저 자동 저장 불가', 'unavailable');
      }} else if (hasStoredComments && input.value) {{
        setCommentStatus(input, '임시저장 복원됨', 'restored');
      }} else {{
        setCommentStatus(input, '입력 시 자동 저장', '');
      }}
      input.addEventListener('input', () => {{
        collectComments();
        setCommentStatus(input, '저장됨', 'saved');
      }});
    }}
    syncCommentsData();
    persistComments();

    const embeddedRoleDescriptions = readEmbeddedRoleDescriptions();
    const storedRoleDescriptions = readStoredRoleDescriptions();
    const hasStoredRoleDescriptions = Object.keys(storedRoleDescriptions).length > 0;
    roleDescriptions = Object.assign({{}}, embeddedRoleDescriptions, storedRoleDescriptions);
    for (const input of roleDescriptionInputs) {{
      input.value = roleDescriptions[input.dataset.role] || '';
      updateRoleDescriptionPrint(input);
      if (!commentStorageAvailable) {{
        setRoleDescriptionStatus(input, '브라우저 자동 저장 불가', 'unavailable');
      }} else if (hasStoredRoleDescriptions && input.value) {{
        setRoleDescriptionStatus(input, '임시저장 복원됨', 'restored');
      }} else {{
        setRoleDescriptionStatus(input, '입력 시 자동 저장', '');
      }}
      input.addEventListener('input', () => {{
        collectRoleDescriptions();
        setRoleDescriptionStatus(input, '저장됨', 'saved');
      }});
    }}
    syncRoleDescriptionsData();
    persistRoleDescriptions();

    document.querySelector('#save-commented-html').addEventListener('click', saveCommentedHtml);
    document.querySelector('#print-pdf').addEventListener('click', () => {{
      preparePrint();
      window.print();
    }});
    window.addEventListener('beforeprint', preparePrint);
    window.addEventListener('afterprint', restoreScreenView);
  </script>
</body>
</html>
"""
    path.write_text(html, encoding="utf-8")


@lru_cache(maxsize=1)
def _html2canvas_source() -> str:
    """Return the vendored browser renderer without requiring a network connection."""

    try:
        source = (
            resource_files("wlc_role_acl_collector")
            .joinpath("static")
            .joinpath("html2canvas.min.js")
            .read_text(encoding="utf-8")
        )
    except (FileNotFoundError, OSError):
        # Report creation must still work if a packaging mistake omits the optional renderer.
        # The generated page will show a clear PNG-export error instead of losing the report.
        return ""
    return source.replace("</script", r"<\/script")


def _role_image_export_script() -> str:
    """Build the standalone-browser logic for exporting the selected Role panel."""

    return r"""
    const roleImageButton = document.querySelector('#save-role-png');
    const roleImageStatus = document.querySelector('#role-image-status');
    const ROLE_IMAGE_SCALE = 2;
    const ROLE_IMAGE_MIN_WIDTH = 1200;
    const ROLE_IMAGE_MAX_WIDTH = 1800;
    const ROLE_IMAGE_MAX_SINGLE_HEIGHT = 12000;
    const ROLE_IMAGE_PAGE_HEIGHT = 5000;

    function setRoleImageStatus(text, state = '') {
      if (!roleImageStatus) {
        return;
      }
      roleImageStatus.textContent = text;
      roleImageStatus.dataset.state = state;
    }

    function roleImageFileName(roleName) {
      const safeRole = String(roleName || 'role')
        .replace(/[<>:"/\\|?*\u0000-\u001f]/g, '_')
        .replace(/\s+/g, '_')
        .replace(/[. ]+$/g, '')
        .slice(0, 80) || 'role';
      const now = new Date();
      const pad = (value) => String(value).padStart(2, '0');
      const stamp = [
        now.getFullYear(),
        pad(now.getMonth() + 1),
        pad(now.getDate()),
        '_',
        pad(now.getHours()),
        pad(now.getMinutes()),
        pad(now.getSeconds()),
      ].join('');
      return `role_${safeRole}_${stamp}`;
    }

    function replaceExportTextarea(cloneInput, sourceInput, className, emptyText = '') {
      const value = sourceInput?.value || emptyText;
      const output = document.createElement('div');
      output.className = className;
      output.textContent = value;
      cloneInput.replaceWith(output);
    }

    function prepareRoleImageClone(panel, exportWidth) {
      const clone = panel.cloneNode(true);
      clone.hidden = false;
      clone.removeAttribute('id');
      clone.classList.add('role-image-export');
      clone.style.width = `${exportWidth}px`;
      clone.style.maxWidth = 'none';

      for (const control of clone.querySelectorAll(
        '.acl-filter-actions, .role-description-status, .comment-status, '
        + '.comment-print, .role-description-print'
      )) {
        control.remove();
      }

      for (const button of clone.querySelectorAll('.alias-link')) {
        const text = document.createElement('span');
        text.className = 'alias-link';
        text.textContent = button.textContent || '';
        button.replaceWith(text);
      }

      for (const input of clone.querySelectorAll('.comment-input')) {
        const sourceInput = commentInputs.find(
          (item) => item.dataset.commentId === input.dataset.commentId
        );
        replaceExportTextarea(input, sourceInput, 'comment-export-text');
      }

      for (const input of clone.querySelectorAll('.role-description-input')) {
        const sourceInput = roleDescriptionInputs.find(
          (item) => item.dataset.role === input.dataset.role
        );
        replaceExportTextarea(
          input,
          sourceInput,
          'role-description-export-text',
          '입력된 설명이 없습니다.'
        );
      }
      return clone;
    }

    function mountRoleImageClone(clone, exportWidth) {
      const host = document.createElement('div');
      host.setAttribute('aria-hidden', 'true');
      host.style.background = '#ffffff';
      host.style.left = '0';
      host.style.pointerEvents = 'none';
      host.style.position = 'fixed';
      host.style.top = '0';
      host.style.width = `${exportWidth}px`;
      host.style.zIndex = '-2147483647';
      host.appendChild(clone);
      document.body.appendChild(host);
      return host;
    }

    function pruneHiddenRoleRows(clone) {
      for (const row of clone.querySelectorAll('tbody > tr')) {
        const style = getComputedStyle(row);
        if (row.hidden || style.display === 'none' || style.visibility === 'hidden') {
          row.remove();
        }
      }
    }

    function assignRoleImageGroups(clone) {
      let groupIndex = -1;
      for (const row of clone.querySelectorAll('tbody > tr')) {
        if (row.classList.contains('acl-rule-row')) {
          groupIndex += 1;
        }
        row.dataset.exportGroup = String(Math.max(groupIndex, 0));
      }
      return groupIndex + 1;
    }

    function roleImageChunks(clone, groupCount) {
      if (groupCount <= 0) {
        return [];
      }
      const tbody = clone.querySelector('tbody');
      if (!tbody) {
        return [];
      }
      const fullHeight = Math.ceil(clone.scrollHeight);
      const bodyHeight = Math.ceil(tbody.getBoundingClientRect().height);
      const availableHeight = Math.max(800, ROLE_IMAGE_PAGE_HEIGHT - (fullHeight - bodyHeight));
      const groupHeights = new Array(groupCount).fill(0);
      for (const row of tbody.querySelectorAll(':scope > tr')) {
        const index = Number(row.dataset.exportGroup || 0);
        groupHeights[index] += Math.ceil(row.getBoundingClientRect().height);
      }

      const chunks = [];
      let current = [];
      let currentHeight = 0;
      groupHeights.forEach((height, index) => {
        if (current.length && currentHeight + height > availableHeight) {
          chunks.push(current);
          current = [];
          currentHeight = 0;
        }
        current.push(index);
        currentHeight += height;
      });
      if (current.length) {
        chunks.push(current);
      }
      return chunks;
    }

    function roleImagePageClone(baseClone, groupIndexes, partNumber, partCount) {
      const page = baseClone.cloneNode(true);
      const allowed = new Set(groupIndexes.map(String));
      for (const row of page.querySelectorAll('tbody > tr')) {
        if (!allowed.has(row.dataset.exportGroup || '0')) {
          row.remove();
        }
      }
      const meta = page.querySelector('.acl-section-header span');
      if (meta) {
        const marker = document.createElement('span');
        marker.className = 'image-export-part';
        marker.textContent = `이미지 ${partNumber}/${partCount}`;
        meta.appendChild(marker);
      }
      return page;
    }

    async function renderRoleImage(clone, exportWidth) {
      const width = Math.ceil(Math.min(ROLE_IMAGE_MAX_WIDTH, Math.max(exportWidth, clone.scrollWidth)));
      const height = Math.ceil(clone.scrollHeight);
      return html2canvas(clone, {
        backgroundColor: '#ffffff',
        height,
        logging: false,
        removeContainer: true,
        scale: ROLE_IMAGE_SCALE,
        useCORS: false,
        width,
        windowHeight: height,
        windowWidth: width,
      });
    }

    function downloadRoleCanvas(canvas, fileName) {
      return new Promise((resolve, reject) => {
        canvas.toBlob((blob) => {
          if (!blob) {
            reject(new Error('PNG 데이터 생성에 실패했습니다.'));
            return;
          }
          const link = document.createElement('a');
          const url = URL.createObjectURL(blob);
          link.href = url;
          link.download = fileName;
          document.body.appendChild(link);
          link.click();
          link.remove();
          setTimeout(() => URL.revokeObjectURL(url), 1000);
          resolve();
        }, 'image/png');
      });
    }

    async function saveSelectedRolePng() {
      const panel = rolePanels.find((item) => item.id === selectedRolePanelId && !item.hidden);
      if (!panel) {
        setRoleImageStatus('저장할 Role을 먼저 선택하세요.', 'error');
        return;
      }
      if (typeof html2canvas !== 'function') {
        setRoleImageStatus('PNG 변환 모듈을 찾지 못했습니다. 프로그램을 다시 빌드하거나 업데이트하세요.', 'error');
        return;
      }

      roleImageButton.disabled = true;
      roleImageButton.textContent = 'PNG 생성 중...';
      setRoleImageStatus('선택한 Role 이미지를 준비하고 있습니다.');
      let measurementHost = null;
      try {
        collectComments();
        collectRoleDescriptions();
        const exportWidth = Math.ceil(Math.min(
          ROLE_IMAGE_MAX_WIDTH,
          Math.max(ROLE_IMAGE_MIN_WIDTH, panel.scrollWidth)
        ));
        const baseClone = prepareRoleImageClone(panel, exportWidth);
        measurementHost = mountRoleImageClone(baseClone, exportWidth);
        pruneHiddenRoleRows(baseClone);
        const groupCount = assignRoleImageGroups(baseClone);
        const baseName = roleImageFileName(panel.dataset.role || 'role');
        const fullHeight = Math.ceil(baseClone.scrollHeight);

        if (fullHeight <= ROLE_IMAGE_MAX_SINGLE_HEIGHT || groupCount <= 0) {
          const canvas = await renderRoleImage(baseClone, exportWidth);
          await downloadRoleCanvas(canvas, `${baseName}.png`);
          setRoleImageStatus('선택한 Role PNG를 저장했습니다.', 'success');
          return;
        }

        const chunks = roleImageChunks(baseClone, groupCount);
        measurementHost.remove();
        measurementHost = null;
        for (let index = 0; index < chunks.length; index += 1) {
          const page = roleImagePageClone(baseClone, chunks[index], index + 1, chunks.length);
          const pageHost = mountRoleImageClone(page, exportWidth);
          try {
            const canvas = await renderRoleImage(page, exportWidth);
            const part = String(index + 1).padStart(2, '0');
            const total = String(chunks.length).padStart(2, '0');
            await downloadRoleCanvas(canvas, `${baseName}_part_${part}_of_${total}.png`);
          } finally {
            pageHost.remove();
          }
          await new Promise((resolve) => setTimeout(resolve, 150));
        }
        setRoleImageStatus(
          `${chunks.length}개 PNG로 나누어 저장했습니다. 필요하면 브라우저에서 여러 파일 다운로드를 허용하세요.`,
          'success'
        );
      } catch (error) {
        setRoleImageStatus(`PNG 생성 실패: ${error?.message || error}`, 'error');
      } finally {
        measurementHost?.remove();
        roleImageButton.disabled = false;
        roleImageButton.textContent = '선택 Role PNG 저장';
      }
    }

    if (roleImageButton) {
      roleImageButton.disabled = rolePanels.length === 0;
      roleImageButton.addEventListener('click', saveSelectedRolePng);
    }
    """


def _json_for_html(data: dict[str, Any]) -> str:
    return (
        json.dumps(data, ensure_ascii=False)
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )


def _access_check_history_data_html(enabled: bool) -> str:
    if not enabled:
        return ""
    return '<textarea id="access-check-history-data" hidden>[]</textarea>'


def _internal_role_network_banner_html(rows: list[dict[str, Any]], *, enabled: bool) -> str:
    if not enabled:
        return ""
    role_names = {
        str(row.get("role", "")).strip().casefold()
        for row in rows
        if str(row.get("role", "")).strip()
    }
    network_count = len(
        [
            row
            for row in rows
            if str(row.get("local_role_network", "")).strip()
        ]
    )
    status_counts: dict[str, int] = {}
    counted_status_roles: set[tuple[str, str]] = set()
    for row in rows:
        role = str(row.get("role", "")).strip().casefold()
        status = str(row.get("status", "")).strip() or "Unknown"
        key = (role, status)
        if key in counted_status_roles:
            continue
        counted_status_roles.add(key)
        status_counts[status] = status_counts.get(status, 0) + 1
    status_items = "".join(
        f"<li>{escape(status)} {count}</li>"
        for status, count in sorted(status_counts.items(), key=lambda item: item[0].casefold())
    )
    return f"""
    <section class="internal-network-banner" aria-label="Internal Role network notice">
      <strong>내부망 전용 보고서</strong>
      <p>사내 Role 대역표가 포함되어 실제 네트워크 대역과 WLC 수집값 비교 상태를 표시합니다. 회사 외부 공유 전 반드시 내용을 확인하세요.</p>
      <ul>
        <li>Role {len(role_names)}개</li>
        <li>대역 {network_count}개</li>
        {status_items}
      </ul>
    </section>
    """


def _executive_summary_html(
    frames: dict[str, pd.DataFrame],
    role_items: list[dict[str, Any]],
    local_network_rows: list[dict[str, Any]],
    *,
    unresolved_count: int,
) -> str:
    ssid_rows = frames.get("SSID_Role_Map", pd.DataFrame()).to_dict(orient="records")
    dynamic_roles = {
        str(row.get("role", "")).strip()
        for row in ssid_rows
        if str(row.get("role", "")).strip() and _bool_value(row.get("dynamic_role_possible"))
    }
    local_attention_statuses = {"Mismatch", "Local mapping missing", "Role not collected"}
    local_attention_roles = {
        (
            str(row.get("role", "")).strip().casefold(),
            str(row.get("status", "")).strip(),
        )
        for row in local_network_rows
        if str(row.get("role", "")).strip()
        and str(row.get("status", "")).strip() in local_attention_statuses
    }
    attention_count = unresolved_count + len(dynamic_roles) + len(local_attention_roles)
    top_roles = [
        item
        for item in role_items[:3]
        if str(item.get("role", "")).strip()
    ]
    if top_roles:
        top_role_html = "".join(
            "<li>"
            f"{escape(str(item.get('role', '')))} "
            f"<span>{_int_value(item.get('user_count'))} users</span>"
            "</li>"
            for item in top_roles
        )
    else:
        top_role_html = "<li>확인 가능한 Role 없음</li>"

    dynamic_text = (
        f"{len(dynamic_roles)}개 Role에서 동적 Role 가능성이 표시되었습니다."
        if dynamic_roles
        else "현재 보고서 기준 동적 Role 가능성이 표시된 항목은 없습니다."
    )
    local_text = (
        f"사내 Role 대역표 비교 확인 필요 {len(local_attention_roles)}건"
        if local_network_rows
        else "사내 Role 대역표를 선택하지 않아 내부 대역 비교는 생략되었습니다."
    )
    conclusion_text = (
        "확인 필요 항목이 있습니다. 아래 Role ACL Detail을 먼저 확인하세요."
        if attention_count
        else "즉시 확인할 고위험 요약 항목은 없습니다. 세부 ACL은 아래 Role ACL Detail에서 확인하세요."
    )

    return f"""
    <section class="executive-summary" aria-label="Report conclusion summary">
      <div class="executive-summary-header">
        <div>
          <h2>결론 요약</h2>
          <p>{escape(conclusion_text)}</p>
        </div>
        <span class="attention-badge">확인 필요 {attention_count}건</span>
      </div>
      <div class="summary-grid">
        <div class="summary-item">
          <strong>Unresolved {unresolved_count}건</strong>
          <span>Alias, ACL, Role 해석이 불완전한 항목입니다.</span>
        </div>
        <div class="summary-item">
          <strong>사용자 많은 Role TOP 3</strong>
          <ul>{top_role_html}</ul>
        </div>
        <div class="summary-item">
          <strong>동적 Role 가능성</strong>
          <span>{escape(dynamic_text)}</span>
        </div>
        <div class="summary-item">
          <strong>사내 Role 대역 비교</strong>
          <span>{escape(local_text)}</span>
        </div>
        <div class="summary-item">
          <strong>Access Check 판정 제한 있음</strong>
          <span>선택한 Role 이름과 정확히 같은 ACL만 보조 판정합니다. 실제 정책 검토는 Role ACL Detail을 기준으로 확인하세요.</span>
        </div>
      </div>
    </section>
    """


def _access_check_controls_html(
    access_check_data: dict[str, Any],
    *,
    history_enabled: bool = False,
    local_role_networks_enabled: bool = False,
) -> str:
    roles = access_check_data.get("roles", [])
    role_options = "".join(
        f"""
        <option value="{escape(str(role.get('role', '')))}">
          {escape(str(role.get('role', '')))} ({escape(str(role.get('userCount', 0)))} users)
        </option>
        """
        for role in roles
    )
    service_options = "".join(
        f'<option value="{escape(str(service))}">{escape(str(service))}</option>'
        for service in access_check_data.get("services", [])
    )
    disabled = " disabled" if not roles else ""
    history_html = ""
    if history_enabled:
        history_html = f"""
      <div class="access-history">
        <div class="access-history-header">
          <h3>Access Check History</h3>
          <button id="clear-access-history" class="report-action secondary" type="button"{disabled}>이력 지우기</button>
        </div>
        <div id="access-check-history" class="access-history-body">
          <span class="notice">Access Check 이력이 없습니다.</span>
        </div>
      </div>
        """
    return f"""
    <section class="access-check no-print" aria-label="Role access check">
      <div class="access-check-title">
        <h2>Access Check</h2>
      </div>
      <div class="access-check-limits">
        <strong>검사 기준</strong>
        <span>보고서에 포함된 ACL/Alias 데이터만 사용하며, 선택한 Role 이름과 정확히 같은 ACL만 판정합니다.</span>
        <span>Service object의 실제 TCP/UDP 포트 내부까지는 해석하지 않습니다.</span>
        {_access_check_local_network_note(local_role_networks_enabled)}
      </div>
      <div class="access-check-grid">
        <label class="access-field">
          <span>Role</span>
          <select id="access-check-role"{disabled}>{role_options}</select>
        </label>
        <label class="access-field">
          <span>Source IP</span>
          <input id="access-check-source" type="text" inputmode="decimal" placeholder="10.10.10.10"{disabled}>
        </label>
        <label class="access-field">
          <span>Destination IP</span>
          <input id="access-check-destination" type="text" inputmode="decimal" placeholder="10.20.20.20"{disabled}>
        </label>
        <label class="access-field">
          <span>Service</span>
          <select id="access-check-service"{disabled}>
            <option value="">자동 - Source/Destination 기준</option>
            {service_options}
          </select>
        </label>
        <button id="run-access-check" class="report-action access-run" type="button"{disabled}>검사</button>
      </div>
      <div id="access-check-result" class="access-check-result" data-status="empty" aria-live="polite">
        <span>검사 결과 없음.</span>
      </div>
      {history_html}
    </section>
    """


def _access_check_local_network_note(enabled: bool) -> str:
    if not enabled:
        return ""
    return "<span>사내 Role 대역표 기준으로 Source IP가 해당 Role 대역 밖이면 경고를 함께 표시합니다.</span>"


def _access_check_css() -> str:
    return """
    .access-check {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      margin: 18px 0;
      overflow: hidden;
      padding: 0;
    }
    .access-check-title {
      align-items: center;
      background: linear-gradient(180deg, #ffffff, #f7fbff);
      border-bottom: 1px solid var(--line);
      display: flex;
      justify-content: space-between;
      padding: 13px 15px;
    }
    .access-check-title h2 {
      font-size: 18px;
      margin: 0;
    }
    .access-check-limits {
      background: #fffaeb;
      border-bottom: 1px solid #fedf89;
      color: #7a2e0e;
      display: grid;
      gap: 4px;
      font-size: 12px;
      padding: 10px 15px;
    }
    .access-check-limits strong {
      font-size: 12px;
    }
    .access-check-grid {
      align-items: end;
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
      padding: 14px 15px;
    }
    .access-field {
      display: flex;
      flex-direction: column;
      gap: 5px;
      min-width: 0;
    }
    .access-field span {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }
    .access-field input,
    .access-field select {
      border: 1px solid var(--line);
      border-radius: 6px;
      color: var(--text);
      font: inherit;
      min-height: 38px;
      padding: 8px 10px;
      width: 100%;
    }
    .access-field input:focus,
    .access-field select:focus,
    .comment-input:focus {
      border-color: #5aa9f4;
      box-shadow: 0 0 0 3px rgba(90, 169, 244, 0.18);
      outline: 0;
    }
    .access-run {
      min-height: 38px;
      width: 100%;
    }
    .access-check-result {
      border: 1px solid var(--line);
      border-radius: 8px;
      margin: 0 15px 14px;
      padding: 10px 12px;
    }
    .access-check-result[data-status="empty"] {
      color: var(--muted);
    }
    .access-check-result[data-status="allowed"] {
      background: #ecfdf3;
      border-color: #abefc6;
    }
    .access-check-result[data-status="blocked"] {
      background: #fef3f2;
      border-color: #fecdca;
    }
    .access-check-result[data-status="special"],
    .access-check-result[data-status="unknown"] {
      background: #fffaeb;
      border-color: #fedf89;
    }
    .access-check-result[data-status="error"] {
      background: #fef3f2;
      border-color: #fecdca;
      color: #b42318;
    }
    .access-result-title {
      align-items: center;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      font-weight: 700;
      margin-bottom: 8px;
    }
    .access-result-title span:first-child {
      font-size: 16px;
    }
    .access-conditional {
      background: #f79009;
      border-radius: 999px;
      color: #ffffff;
      font-size: 11px;
      padding: 2px 7px;
    }
    .access-result-meta {
      display: grid;
      gap: 5px 12px;
      grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
      margin-top: 8px;
    }
    .access-result-meta div {
      min-width: 0;
    }
    .access-result-meta strong {
      color: var(--muted);
      display: block;
      font-size: 11px;
      text-transform: uppercase;
    }
    .access-warning-list {
      margin: 9px 0 0;
      padding-left: 18px;
    }
    .access-rule-match {
      background: #fffbeb;
      outline: 2px solid #f79009;
      outline-offset: -2px;
    }
    .access-history {
      border-top: 1px solid var(--line);
      background: var(--panel-subtle);
      padding: 12px 15px 14px;
    }
    .access-history-header {
      align-items: center;
      display: flex;
      gap: 10px;
      justify-content: space-between;
      margin-bottom: 8px;
    }
    .access-history-header h3 {
      font-size: 14px;
      margin: 0;
    }
    .access-history-header .report-action {
      padding: 7px 9px;
    }
    .access-history-body {
      overflow-x: auto;
    }
    .access-history-table th,
    .access-history-table td {
      font-size: 12px;
      white-space: nowrap;
    }
    .access-history-table td:last-child {
      white-space: normal;
      min-width: 180px;
    }
    """


def _access_check_script(*, history_enabled: bool = False) -> str:
    return """
    const accessCheckData = (() => {
      const dataElement = document.getElementById('access-check-data');
      if (!dataElement) {
        return { roles: [], services: [] };
      }
      try {
        return JSON.parse(dataElement.textContent || '{"roles":[],"services":[]}');
      } catch (_error) {
        return { roles: [], services: [] };
      }
    })();
    const accessRoleInput = document.getElementById('access-check-role');
    const accessSourceInput = document.getElementById('access-check-source');
    const accessDestinationInput = document.getElementById('access-check-destination');
    const accessServiceInput = document.getElementById('access-check-service');
    const accessRunButton = document.getElementById('run-access-check');
    const accessResultElement = document.getElementById('access-check-result');
    const accessHistoryDataElement = document.getElementById('access-check-history-data');
    const accessHistoryElement = document.getElementById('access-check-history');
    const accessClearHistoryButton = document.getElementById('clear-access-history');
    const accessHistoryEnabled = __ACCESS_HISTORY_ENABLED__;
    const accessHistoryStorageKey = accessHistoryEnabled ? `wlc-role-acl-access-history:${location.pathname || document.title}` : '';
    const accessHistoryStorageAvailable = accessHistoryEnabled && (() => {
      try {
        localStorage.setItem('wlc-role-acl-access-history-test', '1');
        localStorage.removeItem('wlc-role-acl-access-history-test');
        return true;
      } catch (_error) {
        return false;
      }
    })();
    let accessCheckHistory = [];

    function accessEscapeHtml(value) {
      return String(value ?? '')
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
    }

    function syncAccessRoleSelection(roleName) {
      const targetRole = String(roleName || '').trim().toLowerCase();
      if (!accessRoleInput || !targetRole) {
        return;
      }
      const option = Array.from(accessRoleInput.options || []).find(
        (item) => String(item.value || '').trim().toLowerCase() === targetRole
      );
      if (!option || accessRoleInput.value === option.value) {
        return;
      }
      accessRoleInput.value = option.value;
      accessRoleInput.dispatchEvent(new Event('change', { bubbles: true }));
    }

    function accessIpToNumber(value) {
      const parts = String(value || '').trim().split('.');
      if (parts.length !== 4) {
        throw new Error(`Invalid IPv4 address: ${value}`);
      }
      let number = 0;
      for (const part of parts) {
        if (!/^\\d+$/.test(part)) {
          throw new Error(`Invalid IPv4 address: ${value}`);
        }
        const octet = Number(part);
        if (!Number.isInteger(octet) || octet < 0 || octet > 255) {
          throw new Error(`Invalid IPv4 address: ${value}`);
        }
        number = number * 256 + octet;
      }
      return number;
    }

    function accessUnique(values) {
      return Array.from(new Set(values.filter(Boolean)));
    }

    function accessReadEmbeddedHistory() {
      if (!accessHistoryEnabled) {
        return [];
      }
      try {
        const parsed = JSON.parse(accessHistoryDataElement?.value || accessHistoryDataElement?.textContent || '[]');
        return Array.isArray(parsed) ? parsed : [];
      } catch (_error) {
        return [];
      }
    }

    function accessReadStoredHistory() {
      if (!accessHistoryEnabled || !accessHistoryStorageAvailable) {
        return [];
      }
      try {
        const parsed = JSON.parse(localStorage.getItem(accessHistoryStorageKey) || '[]');
        return Array.isArray(parsed) ? parsed : [];
      } catch (_error) {
        return [];
      }
    }

    function syncAccessHistoryData() {
      if (!accessHistoryEnabled || !accessHistoryDataElement) {
        return;
      }
      const text = JSON.stringify(accessCheckHistory, null, 2);
      accessHistoryDataElement.value = text;
      accessHistoryDataElement.textContent = text;
    }

    function syncAccessHistoryDomValues() {
      if (!accessHistoryEnabled) {
        return;
      }
      syncAccessHistoryData();
    }

    function persistAccessHistory() {
      if (!accessHistoryEnabled || !accessHistoryStorageAvailable) {
        return;
      }
      localStorage.setItem(accessHistoryStorageKey, JSON.stringify(accessCheckHistory));
    }

    function accessHistoryRecordKey(record) {
      return [
        record.timestamp,
        record.role,
        record.sourceIp,
        record.destinationIp,
        record.service,
        record.verdict,
        record.acl,
        record.sequence,
      ].join('|');
    }

    function accessSetHistory(records) {
      if (!accessHistoryEnabled) {
        return;
      }
      const deduped = [];
      const seen = new Set();
      for (const record of records || []) {
        const key = accessHistoryRecordKey(record || {});
        if (seen.has(key)) {
          continue;
        }
        seen.add(key);
        deduped.push(record);
      }
      accessCheckHistory = deduped.slice(0, 50);
      syncAccessHistoryData();
      persistAccessHistory();
      renderAccessHistory();
    }

    function accessAddHistoryFromResult(result, context) {
      if (!accessHistoryEnabled) {
        return;
      }
      if (!result || result.status === 'error') {
        return;
      }
      const rule = result.matchedRule || {};
      accessSetHistory([
        {
          timestamp: new Date().toISOString(),
          role: context.roleName,
          sourceIp: context.sourceText,
          destinationIp: context.destinationText,
          service: context.selectedService || '(not selected)',
          verdict: result.verdict || '',
          conditional: Boolean(result.conditional),
          acl: rule.acl || '',
          sequence: rule.sequence || '',
          action: rule.action || '',
          ruleId: rule.id || '',
          raw: rule.raw || '',
        },
        ...accessCheckHistory,
      ]);
    }

    function renderAccessHistory() {
      if (!accessHistoryEnabled || !accessHistoryElement) {
        return;
      }
      if (!accessCheckHistory.length) {
        accessHistoryElement.innerHTML = '<span class="notice">Access Check 이력이 없습니다.</span>';
        return;
      }
      const rows = accessCheckHistory.map((record) => `
        <tr>
          <td>${accessEscapeHtml(record.timestamp || '')}</td>
          <td>${accessEscapeHtml(record.role || '')}</td>
          <td>${accessEscapeHtml(record.sourceIp || '')}</td>
          <td>${accessEscapeHtml(record.destinationIp || '')}</td>
          <td>${accessEscapeHtml(record.service || '')}</td>
          <td>${accessEscapeHtml(record.verdict || '')}${record.conditional ? ' (조건부)' : ''}</td>
          <td>${accessEscapeHtml(record.acl || '')}${record.sequence ? ` #${accessEscapeHtml(record.sequence)}` : ''}</td>
          <td>${accessEscapeHtml(record.raw || '')}</td>
        </tr>
      `).join('');
      accessHistoryElement.innerHTML = `
        <table class="access-history-table">
          <thead>
            <tr><th>Time</th><th>Role</th><th>Source</th><th>Destination</th><th>Service</th><th>Verdict</th><th>Rule</th><th>Raw</th></tr>
          </thead>
          <tbody>${rows}</tbody>
        </table>
      `;
    }

    function clearAccessHistory() {
      if (!accessHistoryEnabled) {
        return;
      }
      accessCheckHistory = [];
      syncAccessHistoryData();
      if (accessHistoryStorageAvailable) {
        localStorage.removeItem(accessHistoryStorageKey);
      }
      renderAccessHistory();
    }

    function accessEndpointMatches(ipNumber, matchers, direction, sourceNumber, destinationNumber) {
      let matched = false;
      let uncertain = false;
      const warnings = [];
      for (const matcher of matchers || []) {
        const type = String(matcher.type || '').toLowerCase();
        if (type === 'any') {
          matched = true;
        } else if (type === 'user') {
          matched = matched || direction === 'source' || destinationNumber === sourceNumber;
        } else if (type === 'host' || type === 'network' || type === 'range') {
          const start = Number(matcher.start);
          const end = Number(matcher.end);
          if (Number.isFinite(start) && Number.isFinite(end) && start <= ipNumber && ipNumber <= end) {
            matched = true;
          }
        } else {
          uncertain = true;
          if (matcher.warning) {
            warnings.push(String(matcher.warning));
          }
        }
      }
      return { matched, uncertain, warnings: accessUnique(warnings) };
    }

    function accessServiceMatches(ruleService, selectedService) {
      const normalizedRuleService = String(ruleService || 'any').trim().toLowerCase() || 'any';
      const normalizedSelectedService = String(selectedService || '').trim().toLowerCase();
      if (!normalizedSelectedService) {
        if (normalizedRuleService === 'any') {
          return { matched: true, conditional: false, warnings: [] };
        }
        return {
          matched: true,
          conditional: true,
          warnings: [`Service 자동 모드가 ${ruleService} 전용 rule에 매칭되었습니다. 정확한 Service를 선택해 재확인하세요.`],
        };
      }
      return {
        matched: normalizedRuleService === 'any' || normalizedRuleService === normalizedSelectedService,
        conditional: false,
        warnings: [],
      };
    }

    function accessActionVerdict(action) {
      const normalized = String(action || '').trim().toLowerCase();
      if (normalized === 'deny') {
        return { status: 'blocked', label: '차단(Blocked)' };
      }
      if (normalized === 'permit') {
        return { status: 'allowed', label: '허용(Allowed)' };
      }
      if (['src-nat', 'dst-nat', 'redirect', 'route', 'tunnel', 'forward'].includes(normalized)) {
        return { status: 'special', label: 'NAT/특수 Action 허용' };
      }
      return { status: 'unknown', label: `알 수 없는 action: ${action || 'not parsed'}` };
    }

    function accessLocalWarnings(roleData, sourceNumber, sourceText) {
      const networks = roleData.localNetworks || [];
      if (!networks.length) {
        return [];
      }
      const matched = networks.some((network) => Number(network.start) <= sourceNumber && sourceNumber <= Number(network.end));
      if (matched) {
        return [];
      }
      const labels = networks.map((network) => network.network || network.label).filter(Boolean).join(', ');
      return [`Source IP ${sourceText}가 사내 Role 대역표 범위 밖입니다: ${labels}`];
    }

    function accessClearHighlights() {
      for (const row of document.querySelectorAll('.access-rule-match')) {
        row.classList.remove('access-rule-match');
      }
    }

    function accessHighlightRule(ruleId) {
      if (!ruleId) {
        return;
      }
      const row = Array.from(document.querySelectorAll('[data-rule-id]')).find((item) => item.dataset.ruleId === ruleId);
      if (!row) {
        return;
      }
      const panel = row.closest('.role-panel');
      if (panel) {
        if (panel.classList.contains('zero-user-role') && zeroUserToggleButton && !zeroUserRolesVisible()) {
          zeroUserToggleButton.setAttribute('aria-pressed', 'true');
          syncZeroUserRoles();
        }
        if (typeof selectRolePanel === 'function') {
          selectRolePanel(panel.id);
        }
        if (row.classList.contains('other-acl-rule')) {
          const toggle = panel.querySelector('.toggle-other-acls');
          if (toggle) {
            toggle.setAttribute('aria-pressed', 'true');
            syncOtherAclRows(toggle);
          }
        }
      }
      row.classList.add('access-rule-match');
      row.scrollIntoView({ block: 'center', behavior: 'smooth' });
    }

    function accessRenderResult(result) {
      if (!accessResultElement) {
        return;
      }
      accessResultElement.dataset.status = result.status || 'unknown';
      const warnings = accessUnique(result.warnings || []);
      const warningHtml = warnings.length
        ? `<ul class="access-warning-list">${warnings.map((warning) => `<li>${accessEscapeHtml(warning)}</li>`).join('')}</ul>`
        : '';
      const rule = result.matchedRule;
      const conditional = result.conditional ? '<span class="access-conditional">조건부</span>' : '';
      const ruleHtml = rule
        ? `<div class="access-result-meta">
            <div><strong>ACL</strong>${accessEscapeHtml(rule.acl)}</div>
            <div><strong>Sequence</strong>${accessEscapeHtml(rule.sequence)}</div>
            <div><strong>Action</strong>${accessEscapeHtml(rule.action)}</div>
            <div><strong>Service</strong>${accessEscapeHtml(rule.service || 'any')}</div>
            <div><strong>Source</strong>${accessEscapeHtml(rule.source)}</div>
            <div><strong>Destination</strong>${accessEscapeHtml(rule.destination)}</div>
            <div><strong>Raw</strong>${accessEscapeHtml(rule.raw)}</div>
          </div>`
        : '<div class="access-result-meta"><div><strong>ACL</strong>일치한 ACL rule 없음.</div></div>';
      accessResultElement.innerHTML = `
        <div class="access-result-title">
          <span>${accessEscapeHtml(result.verdict || '')}</span>
          ${conditional}
        </div>
        ${ruleHtml}
        ${warningHtml}
      `;
    }

    function runAccessCheck() {
      accessClearHighlights();
      const roleName = accessRoleInput?.value || '';
      const sourceText = accessSourceInput?.value || '';
      const destinationText = accessDestinationInput?.value || '';
      let sourceNumber;
      let destinationNumber;
      try {
        sourceNumber = accessIpToNumber(sourceText);
        destinationNumber = accessIpToNumber(destinationText);
      } catch (error) {
        accessRenderResult({ status: 'error', verdict: String(error.message || '').replace('Invalid IPv4 address', 'IPv4 주소 형식 오류'), warnings: [] });
        return;
      }
      const roleData = (accessCheckData.roles || []).find((role) => String(role.role || '').toLowerCase() === roleName.toLowerCase());
      if (!roleData) {
        accessRenderResult({ status: 'error', verdict: `Role을 찾을 수 없음: ${roleName}`, warnings: [] });
        return;
      }
      if (!(roleData.rules || []).length) {
        const result = {
          status: 'unknown',
          verdict: '일치하는 Role ACL 없음',
          conditional: false,
          matchedRule: null,
          warnings: ['Access Check는 선택한 Role 이름과 정확히 같은 ACL만 판정합니다.'],
        };
        accessRenderResult(result);
        accessAddHistoryFromResult(result, {
          roleName,
          sourceText,
          destinationText,
          selectedService: accessServiceInput?.value || '',
        });
        return;
      }
      const selectedService = accessServiceInput?.value || '';
      const checkContext = { roleName, sourceText, destinationText, selectedService };
      const localWarnings = accessLocalWarnings(roleData, sourceNumber, sourceText);
      let uncertainCount = 0;
      for (const rule of roleData.rules || []) {
        const sourceResult = accessEndpointMatches(sourceNumber, rule.sourceMatchers, 'source', sourceNumber, destinationNumber);
        const destinationResult = accessEndpointMatches(destinationNumber, rule.destinationMatchers, 'destination', sourceNumber, destinationNumber);
        if (!sourceResult.matched || !destinationResult.matched) {
          if (sourceResult.uncertain || destinationResult.uncertain) {
            uncertainCount += 1;
          }
          continue;
        }
        const serviceResult = accessServiceMatches(rule.service, selectedService);
        if (!serviceResult.matched) {
          continue;
        }
        const verdict = accessActionVerdict(rule.action);
        const warnings = accessUnique([
          ...localWarnings,
          ...sourceResult.warnings,
          ...destinationResult.warnings,
          ...serviceResult.warnings,
          ...(rule.warnings || []),
        ]);
        const result = {
          status: verdict.status,
          verdict: verdict.label,
          conditional: serviceResult.conditional,
          matchedRule: rule,
          warnings,
        };
        accessRenderResult(result);
        accessAddHistoryFromResult(result, checkContext);
        accessHighlightRule(rule.id);
        return;
      }
      const warnings = [...localWarnings];
      if (uncertainCount > 0) {
        warnings.push(`${uncertainCount}개 rule은 alias/name 데이터가 불완전해 완전 판정하지 못했습니다.`);
      }
      const result = {
        status: 'blocked',
        verdict: '기본 차단(Implicit deny)',
        conditional: false,
        matchedRule: null,
        warnings,
      };
      accessRenderResult(result);
      accessAddHistoryFromResult(result, checkContext);
    }

    if (accessRunButton) {
      accessRunButton.addEventListener('click', runAccessCheck);
    }
    if (accessClearHistoryButton) {
      accessClearHistoryButton.addEventListener('click', clearAccessHistory);
    }
    if (accessHistoryEnabled) {
      accessSetHistory([...accessReadEmbeddedHistory(), ...accessReadStoredHistory()]);
    }
    """.replace("__ACCESS_HISTORY_ENABLED__", "true" if history_enabled else "false")


def _acl_rows_html(
    role: str,
    rows: list[dict[str, Any]],
    alias_lookup: dict[str, list[dict[str, Any]]],
) -> str:
    rendered = []
    for index, row in enumerate(rows, start=1):
        is_other_acl = not _is_role_related_acl(role, str(row.get("acl", "")))
        row_class = "acl-rule-row other-acl-rule" if is_other_acl else "acl-rule-row"
        other_acl_attr = ' data-other-acl="true" hidden' if is_other_acl else ""
        rule_id = access_rule_id(role, index)
        comment_id = f"acl-comment-{_safe_dom_id(role)}-{index}"
        detail_id = f"alias-detail-{_safe_dom_id(role)}-{index}"
        source_alias = _alias_name_from_field(str(row.get("source", "")))
        destination_alias = _alias_name_from_field(str(row.get("destination", "")))
        detail_html = _inline_alias_detail_html(
            detail_id=detail_id,
            source_alias=source_alias,
            destination_alias=destination_alias,
            alias_lookup=alias_lookup,
        )
        rendered.append(
            f"""
            <tr class="{row_class}"{other_acl_attr} data-rule-id="{escape(rule_id)}">
              <td>{escape(str(row.get('acl', '')))}</td>
              <td>{escape(str(row.get('sequence', '')))}</td>
              <td>{_action_badge_html(str(row.get('action', '')))}</td>
              <td>{_acl_field_html(str(row.get('source', '')), detail_id, str(row.get('source_interpretation', '')))}</td>
              <td>{_acl_field_html(str(row.get('destination', '')), detail_id, str(row.get('destination_interpretation', '')))}</td>
              <td>{_service_badge_html(str(row.get('service', '')))}</td>
              <td class="raw raw-column">{escape(str(row.get('raw_rule', '')))}</td>
              <td class="comment-cell">
                <textarea class="comment-input" data-comment-id="{escape(comment_id)}" aria-label="ACL comment"></textarea>
                <div id="{escape(comment_id)}-status" class="comment-status">입력 시 자동 저장</div>
                <div id="{escape(comment_id)}-print" class="comment-print"></div>
              </td>
            </tr>
            {detail_html}
            """
        )
    return "".join(rendered)


def _other_acl_toggle_html(panel_id: str, other_acl_count: int) -> str:
    if other_acl_count <= 0:
        return ""
    return f"""
    <div class="acl-filter-actions no-print">
      <button class="acl-filter-button toggle-other-acls" type="button"
        data-panel-id="{escape(panel_id)}" data-other-acl-count="{other_acl_count}" aria-pressed="false">
        Show other ACLs ({other_acl_count})
      </button>
    </div>
    """


def _other_acl_meta_text(other_acl_count: int) -> str:
    if other_acl_count <= 0:
        return ""
    return f" / {other_acl_count} other hidden"


def _action_badge_html(action: str) -> str:
    normalized = action.strip().casefold()
    if normalized == "permit":
        badge_class = "action-permit"
    elif normalized == "deny":
        badge_class = "action-deny"
    elif normalized in {"src-nat", "dst-nat", "redirect", "route", "tunnel", "forward"}:
        badge_class = "action-special"
    else:
        badge_class = "action-unknown"
    label = action.strip() or "unknown"
    return f'<span class="rule-badge {badge_class}">{escape(label)}</span>'


def _service_badge_html(service: str) -> str:
    label = service.strip() or "any"
    return f'<span class="rule-badge service-badge">{escape(label)}</span>'


def _acl_field_html(value: str, detail_id: str, interpretation: str = "") -> str:
    alias = _alias_name_from_field(value)
    if not alias:
        field = escape(value)
    else:
        field = (
        f'<span class="alias-prefix">alias</span>'
        f'<button class="alias-link" type="button" data-detail-id="{escape(detail_id)}" '
        f'aria-expanded="false">{escape(alias)}</button>'
        )
    if not interpretation:
        return field
    return f'{field}<div class="notice">{escape(interpretation)}</div>'


def _inline_alias_detail_html(
    *,
    detail_id: str,
    source_alias: str,
    destination_alias: str,
    alias_lookup: dict[str, list[dict[str, Any]]],
) -> str:
    aliases: list[tuple[str, str]] = []
    if source_alias:
        aliases.append(("Source", source_alias))
    if destination_alias and destination_alias != source_alias:
        aliases.append(("Destination", destination_alias))
    if not aliases:
        return ""

    sections = []
    for position, alias in aliases:
        sections.append(
            f"""
            <div class="alias-detail-title">{escape(position)} alias: {escape(alias)}</div>
            <div>{_alias_chips_html(alias_lookup.get(alias, []))}</div>
            """
        )
    return f"""
    <tr id="{escape(detail_id)}" class="alias-detail-row" hidden>
      <td colspan="8">
        <div class="alias-detail">{''.join(sections)}</div>
      </td>
    </tr>
    """


def _alias_chips_html(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<span class="notice">Alias detail was not collected.</span>'
    return "".join(
        f"""
        <span class="alias-chip">
          <span class="alias-type alias-type-{_alias_type_class(str(row.get('entry_type', '')))}">{escape(_alias_type_label(str(row.get('entry_type', ''))))}</span>
          <span>{escape(str(row.get('value', '')))}</span>
        </span>
        """
        for row in rows
    )


def _alias_type_label(entry_type: str) -> str:
    normalized = entry_type.strip().lower()
    if normalized in {"host", "network", "range", "name"}:
        return normalized.upper()
    return "RAW"


def _alias_type_class(entry_type: str) -> str:
    normalized = entry_type.strip().lower()
    if normalized in {"host", "network", "range", "name"}:
        return normalized
    return "raw"


def _zero_user_role_controls_html(
    *,
    enabled: bool,
    hidden_count: int,
    user_table_counts_reliable: bool,
    role_items: list[dict[str, Any]],
) -> str:
    if enabled and hidden_count > 0:
        return f"""
        <div class="zero-user-role-controls no-print">
          <button id="toggle-zero-user-roles" class="report-action secondary" type="button"
            data-zero-user-role-count="{hidden_count}" aria-pressed="false">
            Show zero-user roles ({hidden_count})
          </button>
        </div>
        """
    if role_items and not user_table_counts_reliable:
        return """
        <p class="zero-user-role-notice">Role user counts were not available from show user-table, so zero-user Role hiding is disabled.</p>
        """
    if role_items and all(_int_value(item.get("user_count")) == 0 for item in role_items):
        return """
        <p class="zero-user-role-notice">All Roles have 0 observed users, so zero-user Role hiding is disabled to avoid an empty report.</p>
        """
    return ""


def _zero_user_role_class(hidden: bool) -> str:
    return " zero-user-role" if hidden else ""


def _zero_user_role_attrs(hidden: bool) -> str:
    return ' data-zero-user-role="true"' if hidden else ""


def _hidden_attr(hidden: bool) -> str:
    return " hidden" if hidden else ""


def _role_description_html(role: str) -> str:
    description_id = f"role-description-{_safe_dom_id(role)}"
    return f"""
    <div class="role-report-description">
      <div class="role-description-header">
        <label for="{escape(description_id)}">보고용 설명</label>
        <span id="{escape(description_id)}-status" class="role-description-status">입력 시 자동 저장</span>
      </div>
      <textarea id="{escape(description_id)}" class="role-description-input"
        data-role="{escape(role)}" data-description-id="{escape(description_id)}"
        aria-label="{escape(role)} Role 보고용 설명"
        placeholder="상급자 보고용 Role 목적, 대상 사용자, 주요 접근 범위와 확인 사항을 입력하세요."></textarea>
      <div id="{escape(description_id)}-print" class="role-description-print">입력된 설명이 없습니다.</div>
    </div>
    """


def _local_role_network_html(rows: list[dict[str, Any]], enabled: bool) -> str:
    if not enabled:
        return ""

    if not rows:
        status = "Local mapping missing"
        notes = "Role was collected from WLC but was not found in the local Role network Excel."
        networks: list[str] = []
    else:
        status = str(rows[0].get("status", "") or "Local mapping loaded")
        notes = str(rows[0].get("notes", ""))
        networks = [
            str(row.get("local_role_network", "")).strip()
            for row in rows
            if str(row.get("local_role_network", "")).strip()
        ]

    network_html = "".join(
        f'<span class="local-subnet-pill"><strong>LOCAL</strong><span>{escape(network)}</span></span>'
        for network in dict.fromkeys(networks)
    )
    if not network_html:
        network_html = '<span class="notice">No local network is defined for this Role.</span>'

    return f"""
    <div class="local-network">
      <div class="local-network-title">Local Role Network</div>
      <span class="local-network-status {_local_network_status_class(status)}">{escape(status)}</span>
      {network_html}
      <div class="local-network-notes">{escape(notes)}</div>
    </div>
    """


def _local_network_status_class(status: str) -> str:
    normalized = status.casefold()
    if "mismatch" in normalized:
        return "status-mismatch"
    if "missing" in normalized:
        return "status-missing"
    if "matched" in normalized:
        return "status-matched"
    if "not collected" in normalized:
        return "status-not-collected"
    return "status-loaded"


def _alias_name_from_field(value: str) -> str:
    stripped = value.strip()
    if stripped.lower().startswith("alias "):
        return stripped.split(" ", 1)[1].strip()
    return ""


def _group_by_alias(alias_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in alias_rows:
        grouped.setdefault(str(row.get("alias", "")), []).append(row)
    return dict(sorted(grouped.items()))


def _group_local_network_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        role = str(row.get("role", "")).strip()
        if role:
            grouped.setdefault(role, []).append(row)
    return dict(sorted(grouped.items()))


def _role_observed_user_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        role = str(row.get("role", ""))
        if not role:
            continue
        counts[role] = max(counts.get(role, 0), _int_value(row.get("observed_user_count", 0)))
    return counts


def _user_table_counts_reliable(rows: list[dict[str, Any]]) -> bool:
    user_table_rows = [
        row for row in rows if str(row.get("command_id", "")).strip() == "user_table"
    ]
    return bool(user_table_rows) and all(_bool_value(row.get("success")) for row in user_table_rows)


def _bool_value(value: Any) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    return str(value).strip().casefold() in {"true", "1", "yes", "y"}


def _other_acl_row_count(role: str, rows: list[dict[str, Any]]) -> int:
    return sum(1 for row in rows if not _is_role_related_acl(role, str(row.get("acl", ""))))


def _is_role_related_acl(role: str, acl_name: str) -> bool:
    role_name = role.strip().casefold()
    acl = acl_name.strip().casefold()
    if not role_name or not acl:
        return False
    return acl == role_name or acl.startswith(role_name) or role_name in acl


def _int_value(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _safe_dom_id(value: str) -> str:
    return "".join(ch if ch.isalnum() else "-" for ch in value).strip("-") or "item"


def _safe_file_name(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in value)
