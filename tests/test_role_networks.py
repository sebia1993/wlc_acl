from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from wlc_role_acl_collector.role_networks import (
    RoleNetworkDefinitionError,
    load_role_network_definitions,
    load_role_network_definitions_with_summary,
)


def test_load_role_network_definitions_normalizes_korean_template(tmp_path):
    path = tmp_path / "role_networks.xlsx"
    _write_workbook(
        path,
        [
            ["Role 이름", "네트워크 대역", "서브넷마스크"],
            ["guest-logon", "10.30.0.5", "255.255.255.0"],
            ["corp-employee", "10.40.1.0/24", ""],
            ["guest-logon", "10.31.0.0", "255.255.0.0"],
        ],
    )

    definitions = load_role_network_definitions(path)

    assert [(item.role, item.network, item.subnet_mask, item.source_row) for item in definitions] == [
        ("guest-logon", "10.30.0.0/24", "255.255.255.0", 2),
        ("corp-employee", "10.40.1.0/24", "255.255.255.0", 3),
        ("guest-logon", "10.31.0.0/16", "255.255.0.0", 4),
    ]


def test_load_role_network_summary_counts_roles_networks_and_duplicates(tmp_path):
    path = tmp_path / "role_networks.xlsx"
    _write_workbook(
        path,
        [
            ["Role 이름", "네트워크 대역", "서브넷마스크"],
            ["guest-logon", "10.30.0.0/24", ""],
            ["guest-logon", "10.30.0.0/24", ""],
            ["corp-employee", "10.40.1.0/24", ""],
        ],
    )

    summary = load_role_network_definitions_with_summary(path)

    assert summary.role_count == 2
    assert summary.network_count == 2
    assert summary.duplicate_count == 1
    assert summary.source_file == str(path)
    assert [(item.role, item.network) for item in summary.definitions] == [
        ("guest-logon", "10.30.0.0/24"),
        ("corp-employee", "10.40.1.0/24"),
    ]


def test_load_role_network_definitions_allows_cidr_without_subnet_mask_column(tmp_path):
    path = tmp_path / "role_networks.xlsx"
    _write_workbook(
        path,
        [
            ["Role 이름", "네트워크 대역"],
            ["guest-logon", "10.30.0.0/24"],
            ["corp-employee", "10.40.1.0/24"],
        ],
    )

    definitions = load_role_network_definitions(path)

    assert [(item.role, item.network, item.subnet_mask) for item in definitions] == [
        ("guest-logon", "10.30.0.0/24", "255.255.255.0"),
        ("corp-employee", "10.40.1.0/24", "255.255.255.0"),
    ]


def test_packaged_role_network_template_contains_input_and_guide_sheets():
    path = Path(__file__).parents[1] / "config" / "role_networks.example.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames[:2] == ["Role_Networks", "작성가이드"]
        input_rows = list(workbook["Role_Networks"].iter_rows(values_only=True))
        guide_text = "\n".join(
            str(value)
            for row in workbook["작성가이드"].iter_rows(values_only=True)
            for value in row
            if value
        )
    finally:
        workbook.close()

    assert input_rows[0][:2] == ("Role 이름", "네트워크 대역")
    assert ("corp-employee", "10.40.2.0/24") == input_rows[3][:2]
    assert "같은 Role에 여러 대역" in guide_text
    assert "내부망 전용" in guide_text

    definitions = load_role_network_definitions(path)
    assert [(item.role, item.network) for item in definitions] == [
        ("guest-logon", "10.30.0.0/24"),
        ("corp-employee", "10.40.1.0/24"),
        ("corp-employee", "10.40.2.0/24"),
    ]


def test_load_role_network_definitions_rejects_invalid_rows(tmp_path):
    path = tmp_path / "role_networks.xlsx"
    _write_workbook(
        path,
        [
            ["role", "network", "subnet_mask"],
            ["guest-logon", "10.30.0.0", ""],
            ["corp-employee", "10.40.1.0/24", "255.255.0.0"],
        ],
    )

    with pytest.raises(RoleNetworkDefinitionError) as exc_info:
        load_role_network_definitions(path)

    message = str(exc_info.value)
    assert "Subnet mask is required" in message
    assert "CIDR prefix and subnet mask do not match" in message


def test_load_role_network_definitions_rejects_renamed_text_file(tmp_path):
    path = tmp_path / "role_networks.xlsx"
    path.write_text("role,network,subnet_mask\ncorp,10.10.10.0,255.255.255.0\n", encoding="utf-8")

    with pytest.raises(RoleNetworkDefinitionError) as exc_info:
        load_role_network_definitions(path)

    message = str(exc_info.value)
    assert "not a valid Excel xlsx/xlsm workbook" in message
    assert "Do not rename a CSV, HTML, or .xls file to .xlsx" in message


def test_load_role_network_definitions_rejects_excel_lock_file(tmp_path):
    path = tmp_path / "~$role_networks.xlsx"
    path.write_bytes(b"")

    with pytest.raises(RoleNetworkDefinitionError) as exc_info:
        load_role_network_definitions(path)

    assert "temporary lock file" in str(exc_info.value)


def test_load_role_network_definitions_rejects_old_xls_file(tmp_path):
    path = tmp_path / "role_networks.xls"
    path.write_bytes(b"old excel bytes")

    with pytest.raises(RoleNetworkDefinitionError) as exc_info:
        load_role_network_definitions(path)

    message = str(exc_info.value)
    assert "old .xls workbook" in message
    assert "save it as Excel Workbook (*.xlsx)" in message


def _write_workbook(path: Path, rows: list[list[str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
